import datetime
import io
import os
import matplotlib.dates as mdates
import matplotlib.patches as patches
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import re

# ==============================================================================
# GELİŞTİRİCİ TANIMLARI
# ==============================================================================
DEVELOPER_NAME = "İlgi Ece Çakmak"

# ==============================================================================
# SAYFA VE GRAFİK YAPILANDIRMASI
# ==============================================================================
st.set_page_config(
    page_title="Sütaş Karacabey Master Scheduler & DSS",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==============================================================================
# KULLANICI & YÖNETİCİ GİRİŞ SİSTEMİ
# ==============================================================================
ADMIN_PIN = "2026"  # 4 Haneli Yönetici Şifresi

def isim_gecerli_mi(isim: str) -> bool:
    isim = isim.strip()
    if not re.fullmatch(r"^[a-zA-ZçÇğĞıİöÖşŞüÜ\s]+$", isim):
        return False
    kelimeler = isim.split()
    if len(kelimeler) < 2:
        return False
    if any(len(k) < 2 for k in kelimeler):
        return False
    return True


def sheet_log_kaydet(kullanici_etiketi: str):
    turkiye_saati = datetime.datetime.utcnow() + datetime.timedelta(hours=3)
    log_time = turkiye_saati.strftime("%d-%m-%Y %H:%M:%S")
    try:
        from streamlit_gsheets import GSheetsConnection
        conn = st.connection("gsheets", type=GSheetsConnection)
        try:
            df_log = conn.read(ttl=0)
            if df_log is None or df_log.empty:
                df_log = pd.DataFrame(columns=["Zaman", "Kullanıcı"])
        except Exception:
            df_log = pd.DataFrame(columns=["Zaman", "Kullanıcı"])
        
        df_log = df_log.dropna(how="all")
        new_row = pd.DataFrame([{"Zaman": log_time, "Kullanıcı": kullanici_etiketi}])
        df_updated = pd.concat([df_log, new_row], ignore_index=True)
        conn.update(data=df_updated)
    except Exception:
        pass


if "auth_user" not in st.session_state:
    st.session_state["auth_user"] = None
if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False
if "admin_login_mode" not in st.session_state:
    st.session_state["admin_login_mode"] = False

if st.session_state["auth_user"] is None:
    # Sayfa ortasında derli toplu giriş kartı
    col_l1, col_l2, col_l3 = st.columns([1, 2, 1])
    
    with col_l2:
        st.markdown("## 🏭 Sütaş Karacabey Scheduler & DSS")
        
        if not st.session_state["admin_login_mode"]:
            st.info("Sisteme erişebilmek için lütfen adınızı ve soyadınızı girip Enter'a basınız.")
            with st.form("login_form", clear_on_submit=False):
                user_name = st.text_input("Ad Soyad:", placeholder="örn: Ahmet Yılmaz")
                submit_btn = st.form_submit_button("Sisteme Giriş Yap ↵", use_container_width=True, type="primary")
                
                if submit_btn:
                    temiz_isim = user_name.strip()
                    if not isim_gecerli_mi(temiz_isim):
                        st.error("⚠️ Lütfen geçerli bir Ad ve Soyad giriniz (Sembol veya rakam kullanılamaz).")
                    else:
                        st.session_state["auth_user"] = temiz_isim.title()
                        st.session_state["is_admin"] = False
                        sheet_log_kaydet(temiz_isim.title())
                        st.rerun()

            st.markdown("---")
            if st.button("👑 Yönetici Olarak Devam Et", use_container_width=True):
                st.session_state["admin_login_mode"] = True
                st.rerun()

        else:
            st.warning("🔒 **Yönetici Giriş Paneli**")
            with st.form("admin_form", clear_on_submit=False):
                admin_name = st.text_input("Yönetici Ad Soyad:", placeholder="örn: Sistem Yöneticisi")
                pin_input = st.text_input("4 Haneli Yönetici Kodunu Giriniz:", type="password", max_chars=4)
                admin_submit = st.form_submit_button("Yetkiyi Doğrula ve Giriş Yap ↵", use_container_width=True, type="primary")
                
                if admin_submit:
                    temiz_isim = admin_name.strip()
                    if not isim_gecerli_mi(temiz_isim):
                        st.error("⚠️ Lütfen geçerli bir Ad ve Soyad giriniz.")
                    elif pin_input != ADMIN_PIN:
                        st.error("❌ Hatalı yönetici kodu! Lütfen tekrar deneyin.")
                    else:
                        st.session_state["auth_user"] = f"{temiz_isim.title()} (Yönetici)"
                        st.session_state["is_admin"] = True
                        st.session_state["admin_login_mode"] = False
                        sheet_log_kaydet(f"{temiz_isim.title()} (Yönetici)")
                        st.rerun()

            if st.button("⬅️ Kullanıcı Girişine Dön", use_container_width=True):
                st.session_state["admin_login_mode"] = False
                st.rerun()

    st.stop()

# ==============================================================================
# MEVCUT UYGULAMA TANIMLARI
# ==============================================================================
plt.rcParams["font.sans-serif"] = "DejaVu Sans"
plt.rcParams["axes.edgecolor"] = "#D9D9D9"
plt.rcParams["axes.linewidth"] = 0.8

MIN_SUT_LIMITI_TON = 0.01
MAKINE_LISTESI = ["Küçük Kova", "Büyük Kova", "132 çap", "160 çap", "Grunwald"]

CIP_HATLARI = {
    "160 çap": "HAT_1",
    "132 çap": "HAT_1",
    "Grunwald": "HAT_1",
    "Küçük Kova": "HAT_2",
    "Büyük Kova": "HAT_2",
}

CIP_SURELERI_DK = {
    "160 çap": 60,
    "132 çap": 60,
    "Grunwald": 110,
    "Büyük Kova": 60,
    "Küçük Kova": 60,
}

TANK_KAPASITELERI = {
    "T43": 38.0,
    "T40": 25.0,
    "T41": 25.0,
    "T42": 25.0,
}

TANK_RENKLERI = {
    "T40": {"fill": "E2EFDA", "font": "276A3C"},
    "T41": {"fill": "DDEBF7", "font": "1B4E75"},
    "T42": {"fill": "FFF2CC", "font": "806000"},
    "T43": {"fill": "FCE4D6", "font": "A61C1C"},
}

DEFAULT_FACTORY_DATA = {
    "Pazartesi": [
        ("KAYMAKSIZ YĞR YY 150 G", 8638.97),
        ("KAYMAKSIZ YĞR YY 200 G", 3839.54),
        ("KAYMAKSIZ YĞR 200 G ED", 1920.0),
        ("KAYMAKSIZ YĞR 500 G 49,50t", 19271.0),
        ("KAYMAKSIZ YĞR 600 G", 3469.0),
        ("KAYMAKSIZ YĞR 1000 G", 19271.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 16670.0),
        ("KAYMAK GİBİ KAYMAKSIZ YĞR 1250 G", 25005.0),
        ("KAYMAKSIZ YĞR YY KOVA 5 KG", 2399.71),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 73143.27),
        ("KAYMAKSIZ YĞR 10 KG ED", 14453.0),
    ],
    "Salı": [
        ("KAYMAKSIZ YĞR YY 150 G", 11519.0),
        ("KAYMAKSIZ YĞR YY 200 G", 5759.0),
        ("KAYMAKSIZ YĞR 200 G (95 ÇAP)", 1927.0),
        ("KAYMAKSIZ YĞR 200 G ED", 1920.0),
        ("KAYMAKSIZ YĞR 500 G 49,50t", 19271.0),
        ("KAYMAKSIZ YĞR 600 G", 3469.0),
        ("KAYMAKSIZ YĞR 600 G (D)", 2313.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G", 936.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G (D)", 935.89),
        ("KAYMAKSIZ YĞR 1000 G", 36134.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 24571.0),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 5855.0),
        ("PAKSÜT KAYMAKSZ YĞR KOVA 10 KG", 63008.0),
        ("KAYMAKSIZ YĞR 10 KG ED", 10407.0),
    ],
    "Çarşamba": [
        ("KAYMAKSIZ YĞR 500 G 49,50t", 23607.0),
        ("KAYMAKSIZ YĞR 600 G", 4047.0),
        ("KAYMAKSIZ YĞR 600 G (D)", 2313.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G", 936.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G (D)", 936.0),
        ("KAYMAKSIZ YĞR 1000 G", 21198.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 19271.0),
        ("KAYMAKSIZ YĞR 1500 G", 21680.0),
        ("KAYMAKSIZ YĞR KOVA 2000 G", 28329.0),
        ("KAYMAKSIZ YĞR YY KOVA 5 KG", 3840.0),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 57305.0),
        ("KAYMAKSIZ YĞR 10 KG ED", 4625.0),
    ],
    "Perşembe": [
        ("KAYMAKSIZ YĞR YY 150 G", 8638.97),
        ("KAYMAKSIZ YĞR YY 200 G", 3839.54),
        ("KAYMAKSIZ YĞR 200 G ED", 1920.0),
        ("KAYMAKSIZ YĞR 500 G 49,50t", 19271.0),
        ("KAYMAKSIZ YĞR 600 G", 3469.0),
        ("KAYMAKSIZ YĞR 1000 G", 19271.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 16670.0),
        ("KAYMAK GİBİ KAYMAKSIZ YĞR 1250 G", 25005.0),
        ("KAYMAKSIZ YĞR YY KOVA 5 KG", 2399.71),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 73143.27),
        ("KAYMAKSIZ YĞR 10 KG ED", 14453.0),
    ],
    "Cuma": [
        ("KAYMAKSIZ YĞR YY 150 G", 11519.0),
        ("KAYMAKSIZ YĞR YY 200 G", 5759.0),
        ("KAYMAKSIZ YĞR 200 G (95 ÇAP)", 1927.0),
        ("KAYMAKSIZ YĞR 200 G ED", 1920.0),
        ("KAYMAKSIZ YĞR 500 G 49,50t", 19271.0),
        ("KAYMAKSIZ YĞR 600 G", 3469.0),
        ("KAYMAKSIZ YĞR 600 G (D)", 2313.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G", 936.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G (D)", 935.89),
        ("KAYMAKSIZ YĞR 1000 G", 36134.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 24571.0),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 5855.0),
        ("PAKSÜT KAYMAKSZ YĞR KOVA 10 KG", 63008.0),
        ("KAYMAKSIZ YĞR 10 KG ED", 10407.0),
    ],
    "Cumartesi": [
        ("KAYMAKSIZ YĞR 500 G 49,50t", 23607.0),
        ("KAYMAKSIZ YĞR 600 G", 4047.0),
        ("KAYMAKSIZ YĞR 600 G (D)", 2313.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G", 936.0),
        ("KAYMAKSIZ YĞR LIGHT 650 G (D)", 936.0),
        ("KAYMAKSIZ YĞR 1000 G", 21198.0),
        ("KAYMAKSIZ YĞR 1000 G 87,50t", 19271.0),
        ("KAYMAKSIZ YĞR 1500 G", 21680.0),
        ("KAYMAKSIZ YĞR KOVA 2000 G", 28329.0),
        ("KAYMAKSIZ YĞR YY KOVA 5 KG", 3840.0),
        ("KAYMAKSIZ YĞR YY KOVA 10 KG ED", 57305.0),
        ("KAYMAKSIZ YĞR 10 KG ED", 4625.0),
    ],
}

URUN_KATALOGU = sorted(list({p[0] for day_rows in DEFAULT_FACTORY_DATA.values() for p in day_rows}))


def create_excel_stream_from_dict(factory_dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in factory_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Açıklama", "Süt Karşılığı (Lt)"])
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def default_excel_stream():
    for f_name in ["123123.xlsx", "haftalik_projeksiyon.xlsx", "Sutas_Projeksiyon.xlsx"]:
        if os.path.exists(f_name):
            with open(f_name, "rb") as f:
                return io.BytesIO(f.read())
    return create_excel_stream_from_dict(DEFAULT_FACTORY_DATA)


def hiz_matrisini_yukle():
    return {
        "160 çap": {
            "750g": {"hiz": 3.024, "sut_tipi": "TAM YAĞLI"},
            "1000g": {"hiz": 3.648, "sut_tipi": "TAM YAĞLI"},
            "1250g": {"hiz": 4.08, "sut_tipi": "%5 YAĞLI"},
            "1500g": {"hiz": 4.032, "sut_tipi": "YAĞLI"},
        },
        "132 çap": {
            "500g": {"hiz": 2.457, "sut_tipi": "TAM YAĞLI"},
            "600g": {"hiz": 2.9484, "sut_tipi": "TAM YAĞLI"},
            "650g": {"hiz": 3.1941, "sut_tipi": "YARIM YAĞLI"},
            "750g": {"hiz": 3.6855, "sut_tipi": "%5 YAĞLI"},
        },
        "Grunwald": {
            "95 çap - 200g": {"hiz": 1.632, "sut_tipi": "TAM YAĞLI"},
            "75 çap - 200g (Tam)": {"hiz": 2.1216, "sut_tipi": "TAM YAĞLI"},
            "75 çap - 200g (Yarım)": {"hiz": 2.1216, "sut_tipi": "YARIM YAĞLI"},
            "75 çap - 150g": {"hiz": 1.836, "sut_tipi": "YARIM YAĞLI"},
        },
        "Küçük Kova": {
            "10000g (Tam)": {"hiz": 6.768, "sut_tipi": "TAM YAĞLI"},
            "10000g (Yarım)": {"hiz": 6.768, "sut_tipi": "YARIM YAĞLI"},
            "10000g (Paksüt)": {"hiz": 6.768, "sut_tipi": "PAKSÜT"},
            "5000g": {"hiz": 5.64, "sut_tipi": "YARIM YAĞLI"},
        },
        "Büyük Kova": {
            "2000g": {"hiz": 3.192, "sut_tipi": "YAĞLI"},
            "10000g (Tam)": {"hiz": 5.415, "sut_tipi": "TAM YAĞLI"},
            "10000g (Yarım)": {"hiz": 5.415, "sut_tipi": "YARIM YAĞLI"},
            "10000g (Paksüt)": {"hiz": 5.415, "sut_tipi": "PAKSÜT"},
        },
    }


MAKINE_HIZLARI = hiz_matrisini_yukle()


def isgucu_katsayisi_getir(makine_adi, gramaj_adi):
    if makine_adi == "160 çap":
        return 4.0
    elif makine_adi == "132 çap":
        return 5.0
    elif makine_adi == "Grunwald":
        return 5.0 if "95" in str(gramaj_adi) else 3.0
    elif makine_adi == "Küçük Kova":
        return 5.0
    elif makine_adi == "Büyük Kova":
        return 6.0
    return 4.0


def sut_tipi_ve_gramaj_tespit(urun_adi, sut_tipi_col="", gramaj_col=""):
    u = str(urun_adi).upper()
    st_col = str(sut_tipi_col).upper()
    g_str = str(gramaj_col).strip()
    full = f"{u} {st_col}"

    if "PAK" in full:
        st = "PAKSÜT"
    elif any(x in full for x in ["%5", "5 YAĞLI", "5 YAGLI", "KAYMAK GİBİ", "KAYMAKGİBİ"]) or "1250" in u:
        st = "%5 YAĞLI"
    elif any(x in full.split() for x in ["YY", "YARIM", "Y.YAĞLI", "Y.YAGLI", "LIGHT", "LİGHT"]) or "650" in u:
        st = "YARIM YAĞLI"
    elif "2000" in u or "1500" in u or ("YAĞLI" in st_col and "TAM" not in st_col):
        st = "YAĞLI"
    else:
        st = "TAM YAĞLI"

    if "10000" in u or "10 KG" in u or "10KG" in u or g_str == "10000":
        g = "10000g"
        m = "KOVA_10KG"
    elif "5000" in u or "5 KG" in u or "5KG" in u or g_str == "5000" or "3000" in u or "3 KG" in u or "3kg" in u or g_str == "3000":
        g = "5000g"
        m = "Küçük Kova"
    elif "2000" in u or "2 KG" in u or "2kg" in u or g_str == "2000":
        g = "2000g"
        m = "Büyük Kova"
    elif "1500" in u or g_str == "1500":
        g = "1500g"
        m = "160 çap"
    elif "1250" in u or g_str == "1250":
        g = "1250g"
        m = "160 çap"
    elif "1000" in u or g_str == "1000":
        g = "1000g"
        m = "160 çap"
    elif "750" in u or g_str == "750":
        g = "750g"
        m = "160 çap" if st in ["TAM YAĞLI", "%5 YAĞLI"] else "132 çap"
    elif "650" in u or g_str == "650":
        g = "650g"
        m = "132 çap"
    elif "600" in u or g_str == "600":
        g = "600g"
        m = "132 çap"
    elif "500" in u or g_str == "500":
        g = "500g"
        m = "132 çap"
    elif "200" in u or g_str == "200":
        m = "Grunwald"
        g = "95 çap - 200g" if ("95" in u or "95 ÇAP" in u) else "75 çap - 200g"
    elif "150" in u or g_str == "150" or "125" in u or "4X125" in u:
        g = "75 çap - 150g"
        m = "Grunwald"
    else:
        g = "1000g"
        m = "160 çap"

    return st, g, m


def makine_hizi_getir(makine_adi, gramaj_adi, sut_tipi):
    if makine_adi == "Küçük Kova":
        return 5.64 if gramaj_adi == "5000g" else 6.768
    elif makine_adi == "Büyük Kova":
        return 3.192 if gramaj_adi == "2000g" else 5.415
    elif makine_adi == "160 çap":
        if gramaj_adi == "750g":
            return 3.024
        if gramaj_adi == "1000g":
            return 3.648
        if gramaj_adi == "1250g":
            return 4.08
        if gramaj_adi == "1500g":
            return 4.032
        return 3.648
    elif makine_adi == "132 çap":
        if gramaj_adi == "500g":
            return 2.457
        if gramaj_adi == "600g":
            return 2.9484
        if gramaj_adi == "650g":
            return 3.1941
        if gramaj_adi == "750g":
            return 3.6855
        return 2.9484
    elif makine_adi == "Grunwald":
        if "95" in str(gramaj_adi):
            return 1.632
        if "150" in str(gramaj_adi):
            return 1.836
        return 2.1216
    return 3.5


def sut_tipi_toplam_hiz_getir(sut_tipi, makineler):
    tot = 0.0
    for m in makineler:
        for g, bil in MAKINE_HIZLARI[m].items():
            if bil["sut_tipi"] == sut_tipi:
                tot += bil["hiz"]
                break
    return max(2.5, tot)


def dinamik_projeksiyon_oku(excel_source, sheet_name):
    df = pd.read_excel(excel_source, sheet_name=sheet_name, header=None)
    header_row = 0
    for r_i in range(min(10, len(df))):
        row_vals = [str(x).lower() for x in df.iloc[r_i].values]
        if any("açıklama" in x or "aciklama" in x for x in row_vals):
            header_row = r_i
            break

    df_header = df.iloc[header_row]
    headers = [str(c).strip().replace("\n", " ") for c in df_header.values]

    aciklama_idx = 0
    miktar_idx = 1
    gramaj_idx = None
    sut_tipi_idx = None

    for idx, h in enumerate(headers):
        h_low = h.lower()
        if "açıklama" in h_low or "aciklama" in h_low:
            aciklama_idx = idx
        elif any(k in h_low for k in ["süt karşılığı", "sut karsiligi", "mamül", "mamul", "miktar"]):
            miktar_idx = idx
        elif "gramaj" in h_low:
            gramaj_idx = idx
        elif "süt tipi" in h_low or "sut tipi" in h_low:
            sut_tipi_idx = idx

    df_data = df.iloc[header_row + 1 :].copy()

    siparisler, idx = [], 1
    for _, row in df_data.iterrows():
        aciklama = str(row.iloc[aciklama_idx]).strip()
        if (
            not aciklama
            or aciklama.lower() in ["nan", "none", ""]
            or "toplam" in aciklama.lower()
            or aciklama.upper() in ["YAĞLI", "Y.YAĞLI", "%5 YAĞLI", "PAKSÜT"]
        ):
            continue

        try:
            val = row.iloc[miktar_idx]
            mamul_kg = float(val) if pd.notnull(val) else 0.0
        except Exception:
            mamul_kg = 0.0

        if mamul_kg > 1.0:
            gramaj_user = str(row.iloc[gramaj_idx]).strip() if gramaj_idx is not None else ""
            sut_tipi_user = str(row.iloc[sut_tipi_idx]).strip() if sut_tipi_idx is not None else ""
            st, g, m = sut_tipi_ve_gramaj_tespit(aciklama, sut_tipi_user, gramaj_user)

            siparisler.append({
                "ana_siparis_id": f"ORD-{idx:02d}",
                "ürün_adı": aciklama,
                "süt_tipi": st,
                "gramaj": g,
                "makine_hedef": m,
                "tonaj_ton": mamul_kg / 1000.0,
            })
            idx += 1
    return siparisler


def ardilsik_uretimleri_birlestir(df_schedule):
    if df_schedule.empty:
        return df_schedule
    merged_rows = []
    curr = None
    for _, row in df_schedule.iterrows():
        if curr is None:
            curr = dict(row)
        else:
            same_machine = row["Makine"] == curr["Makine"]
            same_order = row["Sipariş ID"] == curr["Sipariş ID"]
            same_tank = row["Tahsis Tank"] == curr["Tahsis Tank"]
            same_product = row["Ürün Adı"] == curr["Ürün Adı"]
            same_target = row["04:00 Hedefi"] == curr["04:00 Hedefi"]
            is_continuation = row["Başlangıç"] == curr["Bitiş"]
            if same_machine and same_order and same_tank and same_product and same_target and is_continuation:
                curr["Miktar (Ton)"] = round(curr["Miktar (Ton)"] + row["Miktar (Ton)"], 2)
                curr["Bitiş"] = row["Bitiş"]
                if "dt_end" in row:
                    curr["dt_end"] = row["dt_end"]
            else:
                merged_rows.append(curr)
                curr = dict(row)
    if curr is not None:
        merged_rows.append(curr)
    return pd.DataFrame(merged_rows)


def gunluk_tank_hazirligi_v80(
    day_idx,
    day_name,
    gun_baslangic,
    tank_states,
    assigned_types,
    p6_state,
    audit_log_list,
    p6_debi,
    kultur_suresi,
    p6_cip_limit,
    p6_cip_suresi,
):
    tanks = {}
    tank_list = [("T43", 38.0), ("T40", 25.0), ("T41", 25.0), ("T42", 25.0)]

    if day_idx == 1:
        for idx, (tk_name, cap) in enumerate(tank_list):
            st = assigned_types[idx % len(assigned_types)]
            tanks[tk_name] = {
                "kapasite": cap,
                "mevcut_sut": cap,
                "sut_tipi": st,
                "cip_musait_zaman": gun_baslangic - datetime.timedelta(hours=6),
                "dolum_bitis": gun_baslangic - datetime.timedelta(hours=2),
                "kultur_saati": gun_baslangic - datetime.timedelta(hours=kultur_suresi),
                "hazir_saat": gun_baslangic,
                "bosalma_saati": gun_baslangic,
            }
            audit_log_list.append({
                "Gün": f"GÜN {day_idx} ({day_name})",
                "Tank": tk_name,
                "Kapasite (Ton)": cap,
                "Süt Tipi": st,
                "Önceki Gün Boşalma": "-",
                "Tank CIP Bitiş (Hazır)": "-",
                "P6 Dolum Başlangıç": (gun_baslangic - datetime.timedelta(hours=4.0)).strftime("%d-%m %H:%M"),
                "P6 Bitiş (JIT Kültür)": (gun_baslangic - datetime.timedelta(hours=kultur_suresi)).strftime("%d-%m %H:%M"),
                "P6 Dolum Kuyruğu": "0 dk",
                "Mayalanma Bitiş (Hazır)": gun_baslangic.strftime("%d-%m %H:%M"),
                "Sistemsel Durum & Bekleme Analizi": "✅ Hafta başı başlangıç stoğu: 08:00'de kesintisiz hazır başlatıldı.",
            })
        return tanks

    sorted_tanks = sorted(
        tank_list,
        key=lambda item: tank_states.get(item[0], {}).get("cip_musait_zaman", gun_baslangic - datetime.timedelta(hours=6)),
    )
    night_p6 = max(p6_state["musaitlik"], gun_baslangic - datetime.timedelta(hours=10))

    for idx, (tk_name, cap) in enumerate(sorted_tanks):
        st = assigned_types[idx % len(assigned_types)]
        prev_state = tank_states.get(tk_name, {})
        t_bosaldi = prev_state.get("bosalma_saati", gun_baslangic - datetime.timedelta(hours=7))
        t_cip_done = prev_state.get("cip_musait_zaman", gun_baslangic - datetime.timedelta(hours=6))

        t_p6_start = max(t_cip_done, night_p6)
        p6_kuyruk_dk = int((t_p6_start - t_cip_done).total_seconds() / 60)

        cip_p6_notu = ""
        if p6_state["kumulatif_ton"] + cap > p6_cip_limit:
            t_p6_start += datetime.timedelta(hours=p6_cip_suresi)
            p6_state["kumulatif_ton"] = 0.0
            cip_p6_notu = f" (🧼 P6 {int(p6_cip_limit)}T Limit CIP)"

        dolum_h = cap / p6_debi
        t_p6_end = t_p6_start + datetime.timedelta(hours=dolum_h)
        night_p6 = t_p6_end
        p6_state["kumulatif_ton"] += cap

        actual_ready = max(gun_baslangic, t_p6_end + datetime.timedelta(hours=kultur_suresi))
        kultur_bas = actual_ready - datetime.timedelta(hours=kultur_suresi)

        durum_analizi = ""
        if p6_kuyruk_dk > 0:
            durum_analizi = f"⚠️ P6 Hat Kuyruğu: Tank CIP bitişinden itibaren {p6_kuyruk_dk} dk boyunca P6 pastörizatörünün boşa çıkması beklendi."
            if cip_p6_notu:
                durum_analizi += f" + {p6_cip_suresi} Sa P6 Yıkama."
        else:
            durum_analizi = "✅ P6 hemen müsaitti, CIP sonrası kesintisiz doluma başlandı."

        if actual_ready > gun_baslangic:
            gecikme_dk = int((actual_ready - gun_baslangic).total_seconds() / 60)
            durum_analizi += f" 👉 08:00'e yetişemedi ({gecikme_dk} dk gecikme: JIT Kültür {kultur_bas.strftime('%H:%M')} -> Hazır {actual_ready.strftime('%H:%M')})."
        else:
            durum_analizi += f" 👉 08:00 vardiya başlangıcına zamanında yetişti (JIT Kültür: {kultur_bas.strftime('%H:%M')})."

        tanks[tk_name] = {
            "kapasite": cap,
            "mevcut_sut": cap,
            "sut_tipi": st,
            "cip_musait_zaman": t_cip_done,
            "dolum_bitis": t_p6_end,
            "kultur_saati": kultur_bas,
            "hazir_saat": actual_ready,
            "bosalma_saati": actual_ready,
        }

        audit_log_list.append({
            "Gün": f"GÜN {day_idx} ({day_name})",
            "Tank": tk_name,
            "Kapasite (Ton)": cap,
            "Süt Tipi": st,
            "Önceki Gün Boşalma": t_bosaldi.strftime("%d-%m %H:%M"),
            "Tank CIP Bitiş (Hazır)": t_cip_done.strftime("%d-%m %H:%M"),
            "P6 Dolum Başlangıç": t_p6_start.strftime("%d-%m %H:%M"),
            "P6 Bitiş (JIT Kültür)": t_p6_end.strftime("%d-%m %H:%M") + cip_p6_notu,
            "P6 Dolum Kuyruğu": f"{p6_kuyruk_dk} dk" if p6_kuyruk_dk > 0 else "-",
            "Mayalanma Bitiş (Hazır)": actual_ready.strftime("%d-%m %H:%M"),
            "Sistemsel Durum & Bekleme Analizi": durum_analizi,
        })

    p6_state["musaitlik"] = max(night_p6, gun_baslangic)
    return tanks


def vardiya_ekip_ortalamasi_hesapla(machines_dict, gun_baslangic, mesai_saati=20.0):
    gunduz_bas = gun_baslangic
    gunduz_bit = gun_baslangic + datetime.timedelta(hours=min(10.0, mesai_saati))
    gece_bas = gunduz_bit
    gece_bit = gun_baslangic + datetime.timedelta(hours=mesai_saati)

    gunduz_ornekleri = []
    t = gunduz_bas
    while t < gunduz_bit:
        c = sum(1 for m in MAKINE_LISTESI if any(item[0] <= t < item[1] for item in machines_dict[m]["calisma_araliklari"]))
        gunduz_ornekleri.append(c)
        t += datetime.timedelta(minutes=30)

    gece_ornekleri = []
    t = gece_bas
    while t < gece_bit:
        c = sum(1 for m in MAKINE_LISTESI if any(item[0] <= t < item[1] for item in machines_dict[m]["calisma_araliklari"]))
        gece_ornekleri.append(c)
        t += datetime.timedelta(minutes=30)

    avg_g = max(gunduz_ornekleri) if gunduz_ornekleri else 0
    avg_n = max(gece_ornekleri) if gece_ornekleri else 0
    return avg_g, avg_n


def run_scheduler_pipeline(
    excel_source,
    p6_debi,
    kultur_suresi,
    tank_cip_suresi,
    max_kultur_bekleme,
    makine_max_calisma,
    p6_cip_limit,
    p6_cip_suresi,
    gunluk_mesai_saati=20.0,
    opt_mode="Sezgisel JIT (Mevcut)",
    ariza_aktif=False,
    ariza_gun="Pazartesi",
    ariza_makine="160 çap",
    ariza_saat_str="14:00",
    ariza_sure=60,
):
    xls = pd.ExcelFile(excel_source)
    baslangic_gunu = datetime.datetime(2026, 7, 1, 8, 0)
    mesai_h = int(gunluk_mesai_saati)

    gunluk_cizelgeler = {}
    gunluk_eksikler = {}
    gunluk_makine_istatistikleri = {m: 0.0 for m in MAKINE_LISTESI}
    gunluk_sut_istatistikleri = {}
    haftalik_saatlik_is_yuku = {m: [0.0] * mesai_h for m in MAKINE_LISTESI}
    gunluk_saatlik_isgucu = {}

    audit_log_list = []
    all_schedule_rows = []

    tank_states = {
        "T43": {"cip_musait_zaman": baslangic_gunu - datetime.timedelta(hours=6), "bosalma_saati": baslangic_gunu - datetime.timedelta(hours=6)},
        "T40": {"cip_musait_zaman": baslangic_gunu - datetime.timedelta(hours=6), "bosalma_saati": baslangic_gunu - datetime.timedelta(hours=6)},
        "T41": {"cip_musait_zaman": baslangic_gunu - datetime.timedelta(hours=6), "bosalma_saati": baslangic_gunu - datetime.timedelta(hours=6)},
        "T42": {"cip_musait_zaman": baslangic_gunu - datetime.timedelta(hours=6), "bosalma_saati": baslangic_gunu - datetime.timedelta(hours=6)},
    }

    p6_state = {"musaitlik": baslangic_gunu - datetime.timedelta(hours=6), "kumulatif_ton": 0.0}

    oee_raporu = []
    toplam_talep_genel = 0.0
    toplam_gerceklesen_genel = 0.0
    toplam_eksik_genel = 0.0
    toplam_efektif_p6_saati = 0.0

    for day_idx, sheet_name in enumerate(xls.sheet_names, 1):
        gun_baslangic = baslangic_gunu + datetime.timedelta(days=day_idx - 1)
        cutoff_0400 = gun_baslangic + datetime.timedelta(hours=gunluk_mesai_saati)
        gunluk_saatlik_isgucu[sheet_name] = [0.0] * mesai_h

        siparisler = dinamik_projeksiyon_oku(excel_source, sheet_name)
        if not siparisler:
            continue

        demand_by_type = {}
        for s in siparisler:
            st = s["süt_tipi"]
            demand_by_type[st] = demand_by_type.get(st, 0.0) + s["tonaj_ton"]

        sorted_types = sorted(demand_by_type.keys(), key=lambda k: -demand_by_type[k])
        assigned_types = []
        for pref in ["YARIM YAĞLI", "TAM YAĞLI", "%5 YAĞLI", "PAKSÜT"]:
            if pref in demand_by_type and pref not in assigned_types:
                assigned_types.append(pref)

        for st_k in sorted_types:
            if st_k not in assigned_types:
                assigned_types.append(st_k)

        while len(assigned_types) < 4:
            assigned_types.append(sorted_types[0] if sorted_types else "TAM YAĞLI")

        tanks = gunluk_tank_hazirligi_v80(
            day_idx,
            sheet_name,
            gun_baslangic,
            tank_states,
            assigned_types,
            p6_state,
            audit_log_list,
            p6_debi,
            kultur_suresi,
            p6_cip_limit,
            p6_cip_suresi,
        )

        machines = {
            m: {
                "musait_zamani": gun_baslangic,
                "ardisik_calisma_saat": 0.0,
                "gunluk_toplam_calisma": 0.0,
                "calisma_araliklari": [],
            }
            for m in MAKINE_LISTESI
        }

        is_ariza_gunu = ariza_aktif and (sheet_name.lower().strip() == ariza_gun.lower().strip())
        b_start = None
        b_end = None
        if is_ariza_gunu:
            try:
                h_part, m_part = [int(x) for x in ariza_saat_str.split(":")]
                offset_hours = (h_part - 8) + (m_part / 60.0) if h_part >= 8 else (h_part + 16) + (m_part / 60.0)
                b_start = gun_baslangic + datetime.timedelta(hours=offset_hours)
                b_end = b_start + datetime.timedelta(minutes=ariza_sure)
                machines[ariza_makine]["calisma_araliklari"].append((b_start, b_end, "ARIZA ⚠️"))
            except Exception:
                is_ariza_gunu = False

        cip_hatlari_musaitlik = {"HAT_1": gun_baslangic, "HAT_2": gun_baslangic}
        tank_cip_musaitlik = gun_baslangic - datetime.timedelta(hours=6)

        order_pool = []
        for s in siparisler:
            order_pool.append({
                "siparis_id": s["ana_siparis_id"],
                "ana_id": s["ana_siparis_id"],
                "ürün_adı": s["ürün_adı"],
                "süt_tipi": s["süt_tipi"],
                "gramaj": s["gramaj"],
                "makine_hedef": s["makine_hedef"],
                "orijinal_ton": s["tonaj_ton"],
                "rem_ton": s["tonaj_ton"],
            })

        if "Geçiş" in opt_mode:
            order_pool.sort(key=lambda x: (x["süt_tipi"], x["makine_hedef"]))
        elif "Makespan" in opt_mode:
            order_pool.sort(key=lambda x: x["rem_ton"], reverse=True)

        schedule = []

        while any(o["rem_ton"] > 0.01 for o in order_pool):
            candidate_actions = []
            for m_name in MAKINE_LISTESI:
                if is_ariza_gunu and m_name == ariza_makine:
                    if b_start <= machines[m_name]["musait_zamani"] < b_end:
                        machines[m_name]["musait_zamani"] = b_end

                if machines[m_name]["musait_zamani"] >= cutoff_0400:
                    continue

                m_orders = [
                    o for o in order_pool
                    if o["rem_ton"] > 0.01 and (
                        o["makine_hedef"] == m_name or
                        (o["makine_hedef"] == "KOVA_10KG" and m_name in ["Küçük Kova", "Büyük Kova"])
                    )
                ]
                if m_orders:
                    candidate_actions.append((m_name, machines[m_name]["musait_zamani"]))

            if not candidate_actions:
                break

            candidate_actions.sort(key=lambda x: x[1])
            chosen_m_name = candidate_actions[0][0]
            m_info = machines[chosen_m_name]
            current_time = m_info["musait_zamani"]

            active_count = sum(
                1 for m in MAKINE_LISTESI if any(item[0] <= current_time < item[1] for item in machines[m]["calisma_araliklari"])
            )
            max_allowed = 5

            if active_count >= max_allowed:
                future_ends = [
                    item[1] for m in MAKINE_LISTESI for item in machines[m]["calisma_araliklari"] if item[1] > current_time
                ]
                if future_ends:
                    m_info["musait_zamani"] = min(future_ends)
                    continue

            ready_st_list = [
                tv["sut_tipi"] for tk, tv in tanks.items()
                if tv["mevcut_sut"] > MIN_SUT_LIMITI_TON and (current_time - tv["hazir_saat"]).total_seconds() / 3600.0 <= max_kultur_bekleme
            ]

            matching_orders = [
                o for o in order_pool
                if o["rem_ton"] > 0.01 and (
                    o["makine_hedef"] == chosen_m_name or
                    (o["makine_hedef"] == "KOVA_10KG" and chosen_m_name in ["Küçük Kova", "Büyük Kova"])
                ) and o["süt_tipi"] in ready_st_list
            ]

            if not matching_orders:
                matching_orders = [
                    o for o in order_pool
                    if o["rem_ton"] > 0.01 and (
                        o["makine_hedef"] == chosen_m_name or
                        (o["makine_hedef"] == "KOVA_10KG" and chosen_m_name in ["Küçük Kova", "Büyük Kova"])
                    )
                ]

            if not matching_orders:
                m_info["musait_zamani"] += datetime.timedelta(minutes=15)
                continue

            pending_o = matching_orders[0]
            st_req = pending_o["süt_tipi"]
            g_req = pending_o["gramaj"]
            hiz = makine_hizi_getir(chosen_m_name, g_req, st_req)

            p_start = m_info["musait_zamani"]
            cip_notu = ""

            if m_info["ardisik_calisma_saat"] >= makine_max_calisma:
                hat = CIP_HATLARI[chosen_m_name]
                cip_sure_dk = CIP_SURELERI_DK[chosen_m_name]
                cip_baslangic = max(p_start, cip_hatlari_musaitlik[hat])
                cip_bitis = cip_baslangic + datetime.timedelta(minutes=cip_sure_dk)
                cip_hatlari_musaitlik[hat] = cip_bitis

                m_info["ardisik_calisma_saat"] = 0.0
                m_info["calisma_araliklari"].append((cip_baslangic, cip_bitis, "CIP"))
                p_start = cip_bitis
                cip_notu += f" | 🧼 Makine CIP ({hat}: {cip_sure_dk} dk)"

            matching_tanks = [
                (tk, tv) for tk, tv in tanks.items()
                if tv["sut_tipi"] == st_req and tv["mevcut_sut"] > MIN_SUT_LIMITI_TON and (p_start - tv["hazir_saat"]).total_seconds() / 3600.0 <= max_kultur_bekleme
            ]

            if matching_tanks:
                best_t_name, best_t_info = matching_tanks[0]
                p_start = max(p_start, best_t_info["hazir_saat"])
            else:
                sorted_by_empty = sorted(
                    tanks.items(),
                    key=lambda x: (x[1]["mevcut_sut"] > MIN_SUT_LIMITI_TON, x[1]["bosalma_saati"]),
                )
                refill_t_name, refill_info = sorted_by_empty[0]
                t_bosaldi = refill_info["bosalma_saati"]

                t_cip_start = max(t_bosaldi, tank_cip_musaitlik)
                t_cip_end = t_cip_start + datetime.timedelta(hours=tank_cip_suresi)
                tank_cip_musaitlik = t_cip_end
                refill_info["cip_musait_zaman"] = t_cip_end

                t_p6_start_earliest = max(t_cip_end, p6_state["musaitlik"])
                toplam_st_hizi = sut_tipi_toplam_hiz_getir(st_req, MAKINE_LISTESI)

                kalan_mesai_saati = max(
                    0.0,
                    (cutoff_0400 - (t_p6_start_earliest + datetime.timedelta(hours=1.0 + kultur_suresi))).total_seconds() / 3600.0,
                )
                max_uretilebilir = round(kalan_mesai_saati * toplam_st_hizi, 2)
                rem_demand_st = sum(o["rem_ton"] for o in order_pool if o["süt_tipi"] == st_req)

                fill_amount = min(
                    TANK_KAPASITELERI[refill_t_name],
                    round(rem_demand_st, 2),
                    max(0.0, max_uretilebilir),
                )

                if fill_amount <= 1.0:
                    m_info["musait_zamani"] = cutoff_0400
                    continue

                dolum_suresi = fill_amount / p6_debi
                t_p6_start_jit = max(
                    t_p6_start_earliest,
                    p_start - datetime.timedelta(hours=dolum_suresi + kultur_suresi),
                )

                if p6_state["kumulatif_ton"] + fill_amount > p6_cip_limit:
                    t_p6_start_jit = max(t_p6_start_jit, t_cip_end) + datetime.timedelta(hours=p6_cip_suresi)
                    p6_state["kumulatif_ton"] = 0.0
                    cip_notu += f" | 🧼 P6 {int(p6_cip_limit)}T CIP ({p6_cip_suresi} Sa)"

                p6_end = t_p6_start_jit + datetime.timedelta(hours=dolum_suresi)
                p6_state["musaitlik"] = p6_end
                p6_state["kumulatif_ton"] += fill_amount

                kultur_bas = p6_end
                kultur_hazir = kultur_bas + datetime.timedelta(hours=kultur_suresi)

                tanks[refill_t_name]["mevcut_sut"] = fill_amount
                tanks[refill_t_name]["sut_tipi"] = st_req
                tanks[refill_t_name]["dolum_bitis"] = p6_end
                tanks[refill_t_name]["kultur_saati"] = kultur_bas
                tanks[refill_t_name]["hazir_saat"] = kultur_hazir

                best_t_name = refill_t_name
                best_t_info = tanks[refill_t_name]
                p_start = max(p_start, kultur_hazir)
                cip_notu += f" | 🧼 Tank CIP + P6 Dolum ({round(fill_amount,1)}T)"

            if p_start >= cutoff_0400:
                m_info["musait_zamani"] = cutoff_0400
                continue

            chunk_ton = min(pending_o["rem_ton"], best_t_info["mevcut_sut"])
            if chunk_ton <= MIN_SUT_LIMITI_TON:
                pending_o["rem_ton"] = 0
                continue

            p_dur_h = chunk_ton / hiz
            p_end = p_start + datetime.timedelta(hours=p_dur_h)

            if is_ariza_gunu and chosen_m_name == ariza_makine:
                if p_start < b_start and p_end > b_start:
                    p_end = b_start
                    p_dur_h = (p_end - p_start).total_seconds() / 3600.0
                    chunk_ton = round(p_dur_h * hiz, 2)
                    cip_notu += " | ⚠️ ARIZA KESİNTİSİ"

            if p_end > cutoff_0400:
                p_end = cutoff_0400
                p_dur_h = max(0.0, (cutoff_0400 - p_start).total_seconds() / 3600.0)
                chunk_ton = round(p_dur_h * hiz, 2)

            if chunk_ton <= MIN_SUT_LIMITI_TON:
                m_info["musait_zamani"] = cutoff_0400
                continue

            best_t_info["mevcut_sut"] = max(0.0, round(best_t_info["mevcut_sut"] - chunk_ton, 2))
            best_t_info["bosalma_saati"] = max(best_t_info["bosalma_saati"], p_end)
            pending_o["rem_ton"] = max(0.0, round(pending_o["rem_ton"] - chunk_ton, 2))

            machines[chosen_m_name]["musait_zamani"] = p_end
            machines[chosen_m_name]["ardisik_calisma_saat"] += p_dur_h
            machines[chosen_m_name]["gunluk_toplam_calisma"] += p_dur_h
            machines[chosen_m_name]["calisma_araliklari"].append((p_start, p_end, "URETIM"))

            tank_states[best_t_name]["bosalma_saati"] = max(tank_states[best_t_name]["bosalma_saati"], p_end)
            tank_states[best_t_name]["cip_musait_zaman"] = max(
                tank_states[best_t_name]["cip_musait_zaman"],
                p_end + datetime.timedelta(hours=tank_cip_suresi),
            )

            cult_str = best_t_info["kultur_saati"].strftime("%H:%M") if best_t_info["kultur_saati"] else "06:30"
            ready_str = best_t_info["hazir_saat"].strftime("%H:%M") if best_t_info["hazir_saat"] else "08:00"
            hijyen_notu = f"🧪 Kültür: {cult_str} | ✅ Hazır: {ready_str}{cip_notu}"

            satir_verisi = {
                "Sipariş ID": pending_o["siparis_id"],
                "Ürün Adı": pending_o["ürün_adı"],
                "Süt Tipi": st_req,
                "Miktar (Ton)": round(chunk_ton, 2),
                "Tahsis Tank": best_t_name,
                "Makine": chosen_m_name,
                "Kalıp/Gramaj": g_req,
                "Hız (T/Sa)": hiz,
                "Başlangıç": p_start.strftime("%d-%m-%Y %H:%M"),
                "Bitiş": p_end.strftime("%d-%m-%Y %H:%M"),
                "04:00 Hedefi": "✅ UYGUN",
                "Kültür & CIP Hijyen Notu": hijyen_notu,
                "dt_start": p_start,
                "dt_end": p_end,
                "gun_adi": sheet_name,
            }
            schedule.append(satir_verisi)
            all_schedule_rows.append(satir_verisi)

            op_count = isgucu_katsayisi_getir(chosen_m_name, g_req)
            cur_t = p_start
            while cur_t < p_end:
                h_idx = int((cur_t - gun_baslangic).total_seconds() // 3600)
                if 0 <= h_idx < mesai_h and h_idx < len(haftalik_saatlik_is_yuku[chosen_m_name]):
                    next_hour = gun_baslangic + datetime.timedelta(hours=h_idx + 1)
                    work_in_this_hour = (min(p_end, next_hour) - cur_t).total_seconds() / 3600.0
                    haftalik_saatlik_is_yuku[chosen_m_name][h_idx] += round(work_in_this_hour * hiz, 2)
                    gunluk_saatlik_isgucu[sheet_name][h_idx] += op_count * work_in_this_hour
                    cur_t = min(p_end, next_hour)
                else:
                    break

        if is_ariza_gunu:
            bakim_satiri = {
                "Sipariş ID": "ARIZA-01",
                "Ürün Adı": "⚠️ BAKIM / ARIZA DURUŞU",
                "Süt Tipi": "DURUŞ",
                "Miktar (Ton)": 0,
                "Tahsis Tank": "-",
                "Makine": ariza_makine,
                "Kalıp/Gramaj": "-",
                "Hız (T/Sa)": 0,
                "Başlangıç": b_start.strftime("%d-%m-%Y %H:%M"),
                "Bitiş": b_end.strftime("%d-%m-%Y %H:%M"),
                "04:00 Hedefi": "-",
                "Kültür & CIP Hijyen Notu": "⚠️ Planlı/Plansız Hat Kesintisi",
                "dt_start": b_start,
                "dt_end": b_end,
                "gun_adi": sheet_name,
            }
            schedule.append(bakim_satiri)
            all_schedule_rows.append(bakim_satiri)

        unfulfilled_rows = []
        for o in order_pool:
            if o["rem_ton"] > 0.05:
                uretilen = max(0.0, round(o["orijinal_ton"] - o["rem_ton"], 2))
                unfulfilled_rows.append({
                    "Sipariş ID": o["siparis_id"],
                    "Ürün Adı": o["ürün_adı"],
                    "Süt Tipi": o["süt_tipi"],
                    "Hedef Makine": o["makine_hedef"],
                    "Kalıp/Gramaj": o["gramaj"],
                    "Talep Edilen (Ton)": round(o["orijinal_ton"], 2),
                    "Üretilebilen (Ton)": uretilen,
                    "Eksik Kalan (Ton)": round(o["rem_ton"], 2),
                    "Kalan Neden / Durum": "⚠️ 04:00 Mesai Penceresi Doldu / Günlük Kapasite Tavanı",
                })

        total_day_demand = sum(s["tonaj_ton"] for s in siparisler)
        actual_order_count = len(siparisler)

        df_raw = pd.DataFrame(schedule)
        df_merged = ardilsik_uretimleri_birlestir(df_raw)
        gunluk_cizelgeler[f"GÜN {day_idx} ({sheet_name})"] = df_merged
        gunluk_eksikler[f"GÜN {day_idx} ({sheet_name})"] = pd.DataFrame(unfulfilled_rows)

        if not df_merged.empty:
            for m in MAKINE_LISTESI:
                m_ton = df_merged[df_merged["Makine"] == m]["Miktar (Ton)"].sum()
                gunluk_makine_istatistikleri[m] += m_ton
            for st_val in df_merged["Süt Tipi"].unique():
                if st_val != "DURUŞ":
                    st_ton = df_merged[df_merged["Süt Tipi"] == st_val]["Miktar (Ton)"].sum()
                    gunluk_sut_istatistikleri[st_val] = gunluk_sut_istatistikleri.get(st_val, 0.0) + st_ton

        day_realized = df_merged["Miktar (Ton)"].sum() if not df_merged.empty else 0.0
        day_unfulfilled = sum(r["Eksik Kalan (Ton)"] for r in unfulfilled_rows)

        toplam_talep_genel += total_day_demand
        toplam_gerceklesen_genel += day_realized
        toplam_eksik_genel += day_unfulfilled

        p6_day_pumping_hours = day_realized / p6_debi
        p6_cip_count = max(0, int(day_realized // p6_cip_limit))
        p6_cip_hours = p6_cip_count * p6_cip_suresi
        tank_transition_hours = 0.60

        day_efektif_p6_hours = min(gunluk_mesai_saati, p6_day_pumping_hours + p6_cip_hours + tank_transition_hours)
        toplam_efektif_p6_saati += day_efektif_p6_hours
        p6_efektif_doygunluk = min(100.0, (day_efektif_p6_hours / gunluk_mesai_saati) * 100.0)

        gunduz_ekip, gece_ekip = vardiya_ekip_ortalamasi_hesapla(machines, gun_baslangic, mesai_saati=gunluk_mesai_saati)

        oee_raporu.append({
            "gun": f"GÜN {day_idx}",
            "order_count": actual_order_count,
            "p6_oee": round(p6_efektif_doygunluk, 1),
            "demand": total_day_demand,
            "realized": day_realized,
            "unfulfilled": day_unfulfilled,
            "gunduz_ekip": gunduz_ekip,
            "gece_ekip": gece_ekip,
        })

    # KPI Tablosu
    kpi_rows = []
    toplam_gunduz_ekip_list = []
    toplam_gece_ekip_list = []
    toplam_siparis_sayisi_genel = 0

    for idx, (sheet_title, df_s) in enumerate(gunluk_cizelgeler.items()):
        oee_info = oee_raporu[idx]
        total_demand_ton = oee_info["demand"]
        realized_ton = oee_info["realized"]
        unfulfilled_ton = oee_info["unfulfilled"]
        ontime_pct = (realized_ton / max(0.01, total_demand_ton)) * 100

        toplam_siparis_sayisi_genel += oee_info["order_count"]
        toplam_gunduz_ekip_list.append(oee_info["gunduz_ekip"])
        toplam_gece_ekip_list.append(oee_info["gece_ekip"])

        kpi_rows.append({
            "Gün / Üretim Sayfası": sheet_title,
            "Toplam Sipariş": oee_info["order_count"],
            "Talep Tonajı (Ton)": round(total_demand_ton, 2),
            "Gerçekleşen Üretim (Ton)": round(realized_ton, 2),
            "Üretilemeyen / Kalan (Ton)": round(unfulfilled_ton, 2),
            "Efektif Hat Doygunluğu (%)": f"%{oee_info['p6_oee']}",
            "04:00 Hedef Uyum Oranı (%)": f"%{round(ontime_pct, 1)}",
            "08:00 - 18:00 Ekip": f"{oee_info['gunduz_ekip']} Ekip",
            "18:00 - 04:00 Ekip": f"{oee_info['gece_ekip']} Ekip",
        })

    gun_sayisi = len(gunluk_cizelgeler)
    ort_talep = toplam_talep_genel / max(1, gun_sayisi)
    ort_gerceklesen = toplam_gerceklesen_genel / max(1, gun_sayisi)
    ort_eksik = toplam_eksik_genel / max(1, gun_sayisi)

    genel_p6_doygunluk = min(100.0, (toplam_efektif_p6_saati / (gun_sayisi * gunluk_mesai_saati)) * 100.0)
    genel_uyum = (toplam_gerceklesen_genel / max(0.01, toplam_talep_genel)) * 100
    ort_gunduz_ekip = round(sum(toplam_gunduz_ekip_list) / max(1, len(toplam_gunduz_ekip_list)), 1)
    ort_gece_ekip = round(sum(toplam_gece_ekip_list) / max(1, len(toplam_gece_ekip_list)), 1)

    doygunluk_gece = min(100.0, round(92.5 * (10.0 / p6_debi) * (kultur_suresi / 1.5) * (tank_cip_suresi / 1.0), 1))
    doygunluk_p6 = round(genel_p6_doygunluk, 1)
    doygunluk_tanklar = min(100.0, round(74.0 * (ort_gerceklesen / 163.2) * (kultur_suresi / 1.5) * (20.0 / gunluk_mesai_saati), 1))
    doygunluk_makineler = min(100.0, round((toplam_gerceklesen_genel / (340.0 * gun_sayisi * (gunluk_mesai_saati / 20.0))) * 100, 1))
    tonaj_carpan = ort_gerceklesen / 163.2
    doygunluk_cip = min(100.0, round(25.0 * tonaj_carpan * (8.5 / makine_max_calisma) * (20.0 / gunluk_mesai_saati), 1))

    kpi_rows.append({
        "Gün / Üretim Sayfası": "📊 HAFTALIK GENEL ORTALAMA",
        "Toplam Sipariş": f"{toplam_siparis_sayisi_genel} Sipariş (Toplam)",
        "Talep Tonajı (Ton)": f"{round(ort_talep, 2)} Ton/Gün",
        "Gerçekleşen Üretim (Ton)": f"{round(ort_gerceklesen, 2)} Ton/Gün",
        "Üretilemeyen / Kalan (Ton)": f"{round(ort_eksik, 2)} Ton/Gün",
        "Efektif Hat Doygunluğu (%)": f"%{doygunluk_p6}",
        "04:00 Hedef Uyum Oranı (%)": f"%{round(genel_uyum, 1)}",
        "08:00 - 18:00 Ekip": f"{ort_gunduz_ekip} Ekip (Ort)",
        "18:00 - 04:00 Ekip": f"{ort_gece_ekip} Ekip (Ort)",
    })

    df_kpi = pd.DataFrame(kpi_rows)

    # --------------------------------------------------------------------------
    # EXCEL ÇIKTISI HAZIRLIĞI
    # --------------------------------------------------------------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 1. Ham Sipariş Projeksiyonu Sayfası
    ws_inputs = wb.create_sheet(title="📥 HAM SİPARİŞ PROJEKSİYONU")
    ws_inputs.views.sheetView[0].showGridLines = True
    ws_inputs.append(["Gün", "Ürün Açıklaması", "Süt Karşılığı (Lt)", "Tonaj (Ton)"])
    for col_i in range(1, 5):
        cell_h = ws_inputs.cell(row=1, column=col_i)
        cell_h.font = Font(bold=True, color="FFFFFF")
        cell_h.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell_h.alignment = Alignment(horizontal="center", vertical="center")

    curr_r = 2
    for sheet_n in xls.sheet_names:
        day_orders = dinamik_projeksiyon_oku(excel_source, sheet_n)
        for ord_item in day_orders:
            ws_inputs.append([sheet_n, ord_item["ürün_adı"], ord_item["tonaj_ton"] * 1000.0, ord_item["tonaj_ton"]])
            curr_r += 1

    ws_inputs.column_dimensions["A"].width = 16
    ws_inputs.column_dimensions["B"].width = 38
    ws_inputs.column_dimensions["C"].width = 22
    ws_inputs.column_dimensions["D"].width = 16

    # 2. KPI Sayfası
    ws_kpi = wb.create_sheet(title="📊 YÖNETİCİ ÖZETİ (KPI)")
    ws_kpi.views.sheetView[0].showGridLines = True
    ws_kpi.merge_cells("A1:I2")
    t_cell = ws_kpi["A1"]
    t_cell.value = "SÜTAŞ KARACABEY YOĞURT HATTI - AKILLI ÜRETİM & İŞGÜCÜ DASHBOARD'U"
    t_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_kpi.cell(row=3, column=1, value="1. Günlük Gerçekleşen Üretim & Ortalama Performans Göstergeleri").font = Font(bold=True, size=11, color="1F4E78")

    for col_num, h_text in enumerate(df_kpi.columns, 1):
        c = ws_kpi.cell(row=4, column=col_num, value=h_text)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_i, r_data in enumerate(df_kpi.values, 5):
        is_avg_row = r_i == (5 + len(df_kpi) - 1)
        for c_i, val in enumerate(r_data, 1):
            cell = ws_kpi.cell(row=r_i, column=c_i, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_avg_row)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if is_avg_row:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif r_i % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    kpi_col_widths = {"A": 22, "B": 16, "C": 18, "D": 22, "E": 22, "F": 24, "G": 22, "H": 18, "I": 18}
    for col_letter, w_val in kpi_col_widths.items():
        ws_kpi.column_dimensions[col_letter].width = w_val

    # Grafik Çıktısı (Matplotlib)
    r_graph_start = 5 + len(df_kpi) + 2
    fig_kpi, (ax_k1, ax_k2) = plt.subplots(1, 2, figsize=(10.5, 4.2), dpi=200)
    fig_kpi.patch.set_facecolor("#FFFFFF")

    m_names = list(gunluk_makine_istatistikleri.keys())
    m_tons = [round(gunluk_makine_istatistikleri[m], 1) for m in m_names]
    colors_bar = ["#1D4E89", "#2E6BA8", "#1B3B6F", "#3A404A", "#8EAF9D"]
    bars = ax_k1.bar(m_names, m_tons, color=colors_bar, width=0.65)
    ax_k1.set_title("Makine Bazlı Gerçekleşen Üretim Hacmi (Ton)", fontsize=11, fontweight="bold", pad=12)
    ax_k1.set_ylabel("Gerçekleşen Tonaj (Ton)", fontsize=10)
    ax_k1.grid(axis="y", linestyle="--", alpha=0.5)
    ax_k1.tick_params(axis="x", rotation=15, labelsize=9)
    for bar in bars:
        h = bar.get_height()
        ax_k1.text(bar.get_x() + bar.get_width() / 2, h + 3, f"{h}T", ha="center", va="bottom", fontsize=9, fontweight="bold")

    st_labels = list(gunluk_sut_istatistikleri.keys())
    st_vals = list(gunluk_sut_istatistikleri.values())
    colors_pie = ["#2E6BA8", "#8EAF9D", "#D9E1F2", "#1B3B6F", "#FFC000"]
    if st_vals:
        ax_k2.pie(st_vals, labels=st_labels, autopct="%1.1f%%", startangle=140, colors=colors_pie[: len(st_labels)], textprops={"fontsize": 10})
    ax_k2.set_title("Gerçekleşen Süt Tipi Reçete Dağılımı (%)", fontsize=11, fontweight="bold", pad=12)

    plt.tight_layout()
    buf_kpi = io.BytesIO()
    plt.savefig(buf_kpi, format="png", bbox_inches="tight")
    plt.close(fig_kpi)
    buf_kpi.seek(0)
    img_kpi = OpenpyxlImage(buf_kpi)
    img_kpi.width = 920
    img_kpi.height = 360
    ws_kpi.add_image(img_kpi, f"A{r_graph_start}")

    # 3. Denetim Logu (Audit) Sayfası
    ws_audit = wb.create_sheet(title="🔍 TANK & P6 HAZIRLIK LOGU")
    ws_audit.views.sheetView[0].showGridLines = True
    ws_audit.merge_cells("A1:K2")
    a_title = ws_audit["A1"]
    a_title.value = "📋 SÜTAŞ KARACABEY HATTI - TANK DOLUM & P6 PASTÖRİZATÖR DENETİM GÜNLÜĞÜ (AUDIT LOG)"
    a_title.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    a_title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    a_title.alignment = Alignment(horizontal="center", vertical="center")

    df_audit = pd.DataFrame(audit_log_list)
    for col_num, h_text in enumerate(df_audit.columns, 1):
        c = ws_audit.cell(row=5, column=col_num, value=h_text)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, r_vals in enumerate(df_audit.values, 6):
        tank_name = str(r_vals[1])
        t_style = TANK_RENKLERI.get(tank_name, {"fill": "FFFFFF", "font": "000000"})
        tank_fill = PatternFill(start_color=t_style["fill"], end_color=t_style["fill"], fill_type="solid")
        for c_idx, val in enumerate(r_vals, 1):
            cell = ws_audit.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border
            cell.fill = tank_fill
            if c_idx == 2:
                cell.font = Font(name="Calibri", size=10, bold=True, color=t_style["font"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [1, 3, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    audit_col_widths = {"A": 18, "B": 10, "C": 14, "D": 15, "E": 18, "F": 20, "G": 18, "H": 22, "I": 16, "J": 18, "K": 65}
    for col_letter, w_val in audit_col_widths.items():
        ws_audit.column_dimensions[col_letter].width = w_val

    # Darboğaz & Heatmap Grafikleri
    fig_db1, ax_db1 = plt.subplots(figsize=(10.5, 4.0), dpi=200)
    fig_db1.patch.set_facecolor("#FFFFFF")
    stations = ["CIP Yıkama Devreleri (Hat 1 & 2)", "Dolum Makineleri Parkı (5 Hat)", "Mayalama / Kültür Tank Parkı (113T)", f"P6 Pastörizatör ({p6_debi} Ton/Sa)", "Gece Hazırlığı (04:00 - 08:00)"]
    oee_v = [doygunluk_cip, doygunluk_makineler, doygunluk_tanklar, doygunluk_p6, doygunluk_gece]
    colors_db = ["#FFC000", "#70AD47", "#ED7D31", "#C00000", "#C00000"]
    bars_h = ax_db1.barh(stations, oee_v, color=colors_db, height=0.55)
    ax_db1.set_xlim(0, 120)
    ax_db1.set_xlabel("Kapasite Doluluk Oranı (%)", fontsize=10, fontweight="bold")
    ax_db1.set_title("Tesis İçi Sistem Darboğazları & Kapasite Doluluk Oranları", fontsize=11, fontweight="bold", pad=12)
    ax_db1.grid(axis="x", linestyle="--", alpha=0.5)

    for bar, val in zip(bars_h, oee_v):
        durum_str = "(ANA DARBOĞAZ)" if val > 80 else ("(KRİTİK SÜREÇ RİSKİ)" if val > 70 else "(RAHAT / YEDEKLİ)")
        ax_db1.text(val + 2, bar.get_y() + bar.get_height() / 2, f"%{val} {durum_str}", va="center", fontsize=9, fontweight="bold")

    plt.tight_layout()

    fig_hm, ax_hm = plt.subplots(figsize=(12, 3.8), dpi=200)
    fig_hm.patch.set_facecolor("#FFFFFF")
    hm_data = np.array([[round(v / max(1, gun_sayisi), 1) for v in haftalik_saatlik_is_yuku[m]] for m in MAKINE_LISTESI])
    saatler = [f"{8+i:02d}:00" if 8 + i < 24 else f"{8+i-24:02d}:00" for i in range(mesai_h)]
    max_val = np.max(hm_data) if hm_data.size > 0 and np.max(hm_data) > 0 else 5.0

    cax = ax_hm.pcolor(hm_data, cmap="Blues", edgecolors="white", linewidths=1.5, vmin=0, vmax=max_val)
    ax_hm.set_xticks(np.arange(mesai_h) + 0.5)
    ax_hm.set_xticklabels(saatler, rotation=45, ha="right", fontsize=9, fontweight="bold")
    ax_hm.set_yticks(np.arange(len(MAKINE_LISTESI)) + 0.5)
    ax_hm.set_yticklabels(MAKINE_LISTESI, fontsize=10, fontweight="bold")
    ax_hm.invert_yaxis()

    for y in range(len(MAKINE_LISTESI)):
        for x in range(mesai_h):
            val = hm_data[y, x]
            text_color = "white" if val > (max_val * 0.45) else "black"
            ax_hm.text(x + 0.5, y + 0.5, f"{val:.1f}", ha="center", va="center", color=text_color, fontsize=9, fontweight="bold")

    ax_hm.set_title("Haftalık Ortalama Saatlik Üretim Yoğunluğu Isı Haritası (Heatmap - Ton/Sa)", fontsize=11, fontweight="bold", pad=14)
    ax_hm.set_xlabel(f"Günün Saatleri (08:00 Başlangıçlı {mesai_h} Saatlik Mesai Penceresi)", fontsize=10, fontweight="bold", labelpad=8)
    ax_hm.set_ylabel("Üretim Makineleri", fontsize=10, fontweight="bold")
    cbar = fig_hm.colorbar(cax, ax=ax_hm, fraction=0.025, pad=0.03)
    cbar.set_label("Ortalama Üretim Hacmi (Ton/Sa)", fontsize=9, fontweight="bold")
    cbar.outline.set_visible(False)
    for spine in ax_hm.spines.values():
        spine.set_visible(False)
    plt.tight_layout()

    # 4. Günlük Sayfalar
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    unfulfilled_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cip_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    morning_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    unfulfilled_row_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    sabit_genislikler = {"A": 12, "B": 32, "C": 14, "D": 13, "E": 12, "F": 14, "G": 18, "H": 11, "I": 17, "J": 17, "K": 13, "L": 42}

    for sheet_title, df_detail in gunluk_cizelgeler.items():
        ws_d = wb.create_sheet(title=sheet_title)
        ws_d.views.sheetView[0].showGridLines = True
        display_cols = (
            [c for c in df_detail.columns if c not in ["dt_start", "dt_end", "gun_adi"]]
            if not df_detail.empty
            else ["Sipariş ID", "Ürün Adı", "Süt Tipi", "Miktar (Ton)", "Tahsis Tank", "Makine", "Kalıp/Gramaj", "Hız (T/Sa)", "Başlangıç", "Bitiş", "04:00 Hedefi", "Kültür & CIP Hijyen Notu"]
        )

        for col_num, col_name in enumerate(display_cols, 1):
            c = ws_d.cell(row=1, column=col_num, value=col_name)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = 2
        if not df_detail.empty:
            for _, row in df_detail.iterrows():
                row_vals = [row[col_name] for col_name in display_cols]
                has_cip = "🧼" in str(row_vals[-1])
                is_morning = "08:00" in str(row_vals[-1])
                is_ariza = "⚠️" in str(row_vals[-1])
                for c_idx, val in enumerate(row_vals, 1):
                    c = ws_d.cell(row=current_row, column=c_idx, value=val)
                    c.font = Font(name="Calibri", size=10)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center" if c_idx not in [2, len(display_cols)] else "left", vertical="center")
                    if is_ariza:
                        c.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    elif has_cip:
                        c.fill = cip_fill
                    elif is_morning:
                        c.fill = morning_fill
                current_row += 1

        df_unf = gunluk_eksikler.get(sheet_title, pd.DataFrame())
        if not df_unf.empty:
            current_row += 2
            ws_d.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df_unf.columns))
            title_cell = ws_d.cell(row=current_row, column=1)
            title_cell.value = "❌ 04:00'E YETİŞMEYEN / ÜRETİLEMEYEN SİPARİŞLER"
            title_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            title_cell.fill = unfulfilled_header_fill
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            for col_num, col_name in enumerate(df_unf.columns, 1):
                c = ws_d.cell(row=current_row, column=col_num, value=col_name)
                c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="833C0C", end_color="833C0C", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            for _, row in df_unf.iterrows():
                for c_idx, val in enumerate(row.values, 1):
                    c = ws_d.cell(row=current_row, column=c_idx, value=val)
                    c.font = Font(name="Calibri", size=10)
                    c.fill = unfulfilled_row_fill
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center" if c_idx not in [2, len(row.values)] else "left", vertical="center")
                current_row += 1

        for col_letter, width_val in sabit_genislikler.items():
            ws_d.column_dimensions[col_letter].width = width_val

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return {
        "excel_data": excel_buffer,
        "df_kpi": df_kpi,
        "fig_kpi": fig_kpi,
        "fig_db1": fig_db1,
        "fig_hm": fig_hm,
        "all_schedule_rows": all_schedule_rows,
        "gunluk_cizelgeler": gunluk_cizelgeler,
        "gunluk_eksikler": gunluk_eksikler,
        "df_audit": df_audit,
        "genel_uyum": genel_uyum,
        "genel_p6_oee": genel_p6_doygunluk,
        "ort_gerceklesen": ort_gerceklesen,
        "toplam_talep_genel": toplam_talep_genel,
        "toplam_gerceklesen_genel": toplam_gerceklesen_genel,
        "toplam_eksik_genel": toplam_eksik_genel,
        "gunluk_saatlik_isgucu": gunluk_saatlik_isgucu,
        "mesai_h": mesai_h,
        "gun_sayisi": gun_sayisi,
        "p6_debi": p6_debi,
        "kultur_suresi": kultur_suresi,
    }


# ==============================================================================
# STREAMLIT KULLANICI ARAYÜZÜ (ROL BAZLI DSS VE OPTİMİZASYON)
# ==============================================================================
DEFAULT_PARAMS = {
    "p6_debi": 10.0,
    "kultur_suresi": 1.5,
    "max_kultur_bekleme": 6.0,
    "p6_cip_limit": 100.0,
    "p6_cip_suresi": 1.0,
    "mesai_saati": 20.0,
    "tank_cip_suresi": 1.0,
    "makine_max_calisma": 8.5,
}

for k, v in DEFAULT_PARAMS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if "selected_tab" not in st.session_state:
    st.session_state["selected_tab"] = "📊 Yönetici Özeti"

if "custom_factory_data" not in st.session_state:
    st.session_state["custom_factory_data"] = {k: list(v) for k, v in DEFAULT_FACTORY_DATA.items()}


def varsayilana_sifirla():
    for key, val in DEFAULT_PARAMS.items():
        st.session_state[key] = val
    st.session_state["custom_factory_data"] = {k: list(v) for k, v in DEFAULT_FACTORY_DATA.items()}
    st.session_state["results"] = None
    st.session_state["selected_tab"] = "📊 Yönetici Özeti"
    st.rerun()


# ------------------------------------------------------------------------------
# SAYFA ÜSTÜ: BAŞLIK
# ------------------------------------------------------------------------------
st.title("🏭 Sütaş Karacabey Master Scheduler & DSS")
st.markdown("Tesis kapasite sınırlarına, işgücüne ve CIP döngülerine uygun haftalık üretim, çizelgeleme ve karar destek motoru.")

with st.sidebar:
    # 🏠 ANASAYFAYA DÖN / SIFIRLA BUTONU
    if st.button("🏠 Anasayfa & Varsayılana Dön", use_container_width=True, type="secondary"):
        varsayilana_sifirla()

    st.markdown("---")
    st.markdown("## 🏭 Fabrika Kontrol Paneli")
    
    if st.session_state["is_admin"]:
        st.success("👑 **Yönetici Modu Açık**")
        if st.button("🔄 Normal Kullanıcı Moduna Dön", use_container_width=True):
            st.session_state["is_admin"] = False
            st.session_state["auth_user"] = "Misafir Kullanıcı"
            st.rerun()
    else:
        st.markdown(f"👤 **Giriş Yapan:** `{st.session_state['auth_user']}`")
        with st.expander("🔑 **Yönetici Moduna Geç**", expanded=False):
            elevate_pin = st.text_input("4 Haneli Kod:", type="password", max_chars=4, key="elevate_pin_input")
            if st.button("Yetkiyi Yükselt 🔓", key="btn_elevate", use_container_width=True):
                if elevate_pin == ADMIN_PIN:
                    st.session_state["is_admin"] = True
                    st.session_state["auth_user"] = f"{st.session_state['auth_user']} (Yönetici)"
                    sheet_log_kaydet(f"{st.session_state['auth_user']} [Mod Yükseltme]")
                    st.rerun()
                else:
                    st.error("❌ Hatalı PIN!")

    st.markdown("---")

    # 1. VERİ KAYNAĞI
    st.header("📂 1. Veri Kaynağı")
    veri_secenekleri = ["Sütaş Karacabey Haftalık Projeksiyon (Varsayılan)", "📁 Kendi Excel Dosyamı Yükle"]
    if st.session_state["is_admin"]:
        veri_secenekleri.append("✏️ Ham Veri Düzenleme (Yönetici)")

    veri_secenegi = st.radio("Veri Yöntemini Seçin:", veri_secenekleri, index=0)

    active_excel_source = None
    if veri_secenegi == "📁 Kendi Excel Dosyamı Yükle":
        uploaded_file = st.file_uploader("Projeksiyon Excel Dosyası Seçin (.xlsx)", type=["xlsx"])
        if uploaded_file is not None:
            active_excel_source = uploaded_file
    elif veri_secenegi == "✏️ Ham Veri Düzenleme (Yönetici)":
        active_excel_source = create_excel_stream_from_dict(st.session_state["custom_factory_data"])
    else:
        active_excel_source = create_excel_stream_from_dict(st.session_state["custom_factory_data"])
        st.success("✅ Sütaş Karacabey 6 günlük fabrika projeksiyonu aktif.")

    st.markdown("---")

    # 2. SENARYO & PARAMETRE AYARLARI
    if st.session_state["is_admin"]:
        st.header("🎛️ 2. Senaryo & Parametre Ayarları ✏️")
        st.button("🔄 Parametreleri Varsayılana Sıfırla", on_click=varsayilana_sifirla, use_container_width=True)

        with st.expander("⚡ Pastörizatör (P6) & Mayalama", expanded=False):
            sim_p6_debi = st.slider("P6 Debi Hızı (Ton / Saat)", min_value=6.0, max_value=18.0, value=float(st.session_state.get("p6_debi", DEFAULT_PARAMS["p6_debi"])), step=0.5, key="p6_debi")
            sim_kultur_suresi = st.slider("Mayalama (Kültür) Süresi (Saat)", min_value=1.0, max_value=3.0, value=float(st.session_state.get("kultur_suresi", DEFAULT_PARAMS["kultur_suresi"])), step=0.25, key="kultur_suresi")
            sim_max_kultur_bekleme = st.slider("Maks. Mayalı Bekleme Limiti (Saat)", min_value=3.0, max_value=10.0, value=float(st.session_state.get("max_kultur_bekleme", DEFAULT_PARAMS["max_kultur_bekleme"])), step=0.5, key="max_kultur_bekleme")
            sim_p6_cip_limit = st.number_input("P6 CIP Yıkama Limiti (Ton)", min_value=50.0, max_value=200.0, value=float(st.session_state.get("p6_cip_limit", DEFAULT_PARAMS["p6_cip_limit"])), step=10.0, key="p6_cip_limit")
            sim_p6_cip_suresi = st.slider("P6 CIP Yıkama Süresi (Saat)", min_value=0.5, max_value=2.0, value=float(st.session_state.get("p6_cip_suresi", DEFAULT_PARAMS["p6_cip_suresi"])), step=0.25, key="p6_cip_suresi")

        with st.expander("⏱️ Vardiya & Hijyen Süreleri", expanded=False):
            sim_mesai_saati = st.slider("Günlük Mesai Penceresi (Saat)", min_value=16.0, max_value=24.0, value=float(st.session_state.get("mesai_saati", DEFAULT_PARAMS["mesai_saati"])), step=1.0, key="mesai_saati")
            sim_tank_cip_suresi = st.slider("Tank CIP Süresi (Saat)", min_value=0.5, max_value=2.0, value=float(st.session_state.get("tank_cip_suresi", DEFAULT_PARAMS["tank_cip_suresi"])), step=0.25, key="tank_cip_suresi")
            sim_makine_max_calisma = st.slider("Maks. Ardışık Makine Çalışması (Saat)", min_value=4.0, max_value=12.0, value=float(st.session_state.get("makine_max_calisma", DEFAULT_PARAMS["makine_max_calisma"])), step=0.5, key="makine_max_calisma")

        with st.expander("🤖 Çizelgeleme Algoritması (Optimizasyon)", expanded=False):
            sim_opt_mode = st.radio(
                "Algoritma Hedefi:",
                ["Sezgisel JIT (Mevcut)", "Min-Geçiş (CIP Optimizasyonu)", "Min-Makespan (Kapasite Öncelikli)"],
                help="• Sezgisel JIT: Sütü tam vaktinde mayalayarak tank beklemesini sıfırlar.\n• Min-Geçiş: Aynı reçete ve kalıpları ardışık dizerek CIP duruşlarını minimize eder.\n• Min-Makespan: En büyük siparişleri önceleyerek vardiya bitiş süresini erkene çeker."
            )
    else:
        st.header("🎛️ 2. Aktif Fabrika Parametreleri 🔒")
        sim_p6_debi = DEFAULT_PARAMS["p6_debi"]
        sim_kultur_suresi = DEFAULT_PARAMS["kultur_suresi"]
        sim_max_kultur_bekleme = DEFAULT_PARAMS["max_kultur_bekleme"]
        sim_p6_cip_limit = DEFAULT_PARAMS["p6_cip_limit"]
        sim_p6_cip_suresi = DEFAULT_PARAMS["p6_cip_suresi"]
        sim_mesai_saati = DEFAULT_PARAMS["mesai_saati"]
        sim_tank_cip_suresi = DEFAULT_PARAMS["tank_cip_suresi"]
        sim_makine_max_calisma = DEFAULT_PARAMS["makine_max_calisma"]
        sim_opt_mode = "Sezgisel JIT (Mevcut)"

        with st.expander("⚡ Pastörizatör (P6) & Mayalama (Sabit)", expanded=False):
            st.markdown(f"""
            * **P6 Debi Hızı:** `{sim_p6_debi} Ton/Saat`
            * **Mayalama (Kültür) Süresi:** `{sim_kultur_suresi} Saat`
            * **Maks. Mayalı Bekleme Limiti:** `{sim_max_kultur_bekleme} Saat`
            * **P6 CIP Yıkama Limiti:** `{int(sim_p6_cip_limit)} Ton`
            * **P6 CIP Yıkama Süresi:** `{sim_p6_cip_suresi} Saat`
            """)

        with st.expander("⏱️ Vardiya & Hijyen Süreleri (Sabit)", expanded=False):
            st.markdown(f"""
            * **Günlük Mesai Penceresi:** `{int(sim_mesai_saati)} Saat (08:00 - 04:00)`
            * **Tank CIP Süresi:** `{sim_tank_cip_suresi} Saat`
            * **Maks. Ardışık Makine Çalışması:** `{sim_makine_max_calisma} Saat`
            """)

        with st.expander("🤖 Çizelgeleme Algoritması (Sabit)", expanded=False):
            st.markdown(f"""
            * **Optimizasyon Yöntemi:** `{sim_opt_mode}`
            """)

    st.markdown("---")

    # HERKESE AÇIK: Dinamik Arıza Simülasyonu
    st.header("⚙️ Dinamik Simülasyon & Arıza")
    with st.expander("⚠️ Dinamik Arıza Simülasyonu", expanded=True):
        sim_ariza_aktif = st.toggle("🚨 Arıza Simülasyonunu Devreye Al", value=False)
        if sim_ariza_aktif:
            st.caption("📅 **Arıza Günü Seçimi:**")
            gun_listesi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi"]
            sim_ariza_gun = st.segmented_control("Gün", gun_listesi, default="Salı", label_visibility="collapsed") or "Salı"

            st.caption("⚙️ **Arızalanacak Makine:**")
            sim_ariza_makine = st.segmented_control("Makine", MAKINE_LISTESI, default="160 çap", label_visibility="collapsed") or "160 çap"

            col_ar1, col_ar2 = st.columns(2)
            saat_secenekleri = [f"{h:02d}:{m:02d}" for h in range(8, 24) for m in [0, 15, 30, 45]] + [f"{h:02d}:{m:02d}" for h in range(0, 4) for m in [0, 15, 30, 45]]

            with col_ar1:
                sim_ariza_saat_str = st.select_slider("⏰ Başlangıç Saati", options=saat_secenekleri, value="14:30")
            with col_ar2:
                sim_ariza_sure = st.slider("⏱️ Süre (Dakika)", min_value=15, max_value=240, value=60, step=15)
            st.info(f"📍 **Senaryo:** {sim_ariza_gun} günü saat **{sim_ariza_saat_str}**'de **{sim_ariza_makine}** hattında **{sim_ariza_sure} dk** arıza uygulanacak.")
        else:
            sim_ariza_gun, sim_ariza_makine, sim_ariza_saat_str, sim_ariza_sure = "Pazartesi", "160 çap", "14:00", 60

    st.markdown("---")
    # HERKESE AÇIK: Değişmeyen Sabit Tesis Kısıtları
    st.header("🔒 Sabit Tesis & Fiziksel Kısıtlar")
    with st.expander("🛢️ Mayalama Tank Parkı (113 Ton)", expanded=False):
        st.markdown("""
        * **T43:** 38.0 Ton
        * **T40:** 25.0 Ton
        * **T41:** 25.0 Ton
        * **T42:** 25.0 Ton
        * **Toplam Kapasite:** 113.0 Ton
        * **Asgari Parti Kuralı:** Tanklar boşaldığında minimum 25 Ton parti büyüklüğüyle doldurulur (tam parti mayalama prensibi).
        """)

    with st.expander("🧼 Makine CIP Yıkama Hatları", expanded=False):
        st.markdown("""
        * **HAT_1 (Kase Grubu):**
          * **160 çap:** 60 dk yıkama
          * **132 çap:** 60 dk yıkama
          * **Grunwald:** 110 dk yıkama
        * **HAT_2 (Kova Grubu):**
          * **Küçük Kova:** 60 dk yıkama
          * **Büyük Kova:** 60 dk yıkama
        * **Kural:** Aynı CIP hattına bağlı makineler aynı anda yıkamaya giremez. Yıkama kuyruğu sıralı yönetilir.
        """)

    with st.expander("⚡ Makine Hız Matrisi (Nominal)", expanded=False):
        st.markdown("""
        * **Küçük Kova:** 10 KG (6.77 T/Sa) | 5 KG (5.64 T/Sa)
        * **Büyük Kova:** 10 KG (5.42 T/Sa) | 2 KG (3.19 T/Sa)
        * **160 çap:** 1000g (3.65 T/Sa) | 1250g (4.08 T/Sa) | 1500g (4.03 T/Sa)
        * **132 çap:** 500g (2.46 T/Sa) | 600g (2.95 T/Sa) | 650g (3.19 T/Sa)
        * **Grunwald:** 75 çap (2.12 T/Sa) | 95 çap (1.63 T/Sa) | 150g (1.84 T/Sa)
        """)

    with st.expander("👥 Hat & Operatör İşgücü Katsayıları", expanded=False):
        st.markdown("""
        * **160 çap:** 4.0 Operatör/Saat
        * **132 çap:** 5.0 Operatör/Saat
        * **Grunwald:** 3.0 Operatör/Saat (95 çap: 5.0)
        * **Küçük Kova:** 5.0 Operatör/Saat
        * **Büyük Kova:** 6.0 Operatör/Saat
        * **Eşzamanlı Çalışma:** Maks. 5 Hat (Gündüz & Gece)
        """)

    st.markdown("---")
    st.markdown(
        f"""
        <div style="color: #111111; font-weight: bold; font-size: 13px; padding-bottom: 5px;">
            🛡️ © 2026 Sütaş DSS Platformu
        </div>
        <div style="color: #111111; font-weight: bold; font-size: 13px;">
            Developer: <span style="font-weight: 800;">{DEVELOPER_NAME}</span>
        </div>
        """,
        unsafe_allow_html=True
    )

# ==============================================================================
# HAM VERİ DÜZENLEME EKRANI (YÖNETİCİ MODUNDA SEÇİLDİĞİNDE AÇILIR)
# ==============================================================================
if st.session_state["is_admin"] and veri_secenegi == "✏️ Ham Veri Düzenleme (Yönetici)":
    st.subheader("✏️ Fabrika Haftalık Sipariş Projeksiyonunu Düzenle")
    st.markdown("Aşağıdaki listeden gün seçip mevcut siparişleri silebilir, litrelerini doğrudan değiştirebilir veya ürün listesinden yeni sipariş ekleyebilirsiniz:")

    edit_day = st.selectbox("📅 Düzenlenecek Günü Seçin:", list(st.session_state["custom_factory_data"].keys()))
    current_day_orders = list(st.session_state["custom_factory_data"][edit_day])

    with st.form("custom_data_edit_form"):
        st.write(f"### 📋 {edit_day} Günü Sipariş Listesi")
        
        updated_day_list = []
        col_h1, col_h2, col_h3 = st.columns([5, 3, 2])
        col_h1.markdown("**Ürün Açıklaması**")
        col_h2.markdown("**Süt Miktarı (Lt) ✏️**")
        col_h3.markdown("**Siparişi Sil 🗑️**")

        for idx, (p_name, p_qty) in enumerate(current_day_orders):
            c1, c2, c3 = st.columns([5, 3, 2])
            with c1:
                name_val = st.text_input(f"Ürün #{idx+1}", value=p_name, key=f"name_{edit_day}_{idx}", label_visibility="collapsed")
            with c2:
                qty_val = st.number_input(f"Miktar #{idx+1}", value=float(p_qty), step=100.0, min_value=0.0, key=f"qty_{edit_day}_{idx}", label_visibility="collapsed")
            with c3:
                delete_check = st.checkbox("🗑️ SİL", key=f"del_{edit_day}_{idx}")

            if not delete_check and name_val.strip() != "":
                updated_day_list.append((name_val.strip(), float(qty_val)))

        st.markdown("---")
        st.markdown("**➕ Bu Güne Yeni Sipariş Ekle (Ürün Seçimi):**")
        col_n1, col_n2 = st.columns([5, 3])
        with col_n1:
            new_p_name = st.selectbox("Ürün Seçiniz:", options=["(Ürün Seçin)"] + URUN_KATALOGU, key=f"select_new_p_{edit_day}")
        with col_n2:
            new_p_qty = st.number_input("Süt Miktarı (Lt):", min_value=0.0, step=500.0, value=0.0, key=f"select_new_qty_{edit_day}")

        if new_p_name != "(Ürün Seçin)" and new_p_qty > 0:
            updated_day_list.append((new_p_name.strip(), float(new_p_qty)))

        st.markdown("---")
        confirm_btn = st.form_submit_button("💾 Değişiklikleri Onayla & Üretim Planına Aktar", type="primary", use_container_width=True)

        if confirm_btn:
            st.session_state["custom_factory_data"][edit_day] = updated_day_list
            st.success(f"✅ {edit_day} günü siparişleri başarıyla güncellendi! Aşağıdaki butona basarak yeni planı optimize edebilirsiniz.")
            st.rerun()

    st.markdown("---")

# ==============================================================================
# OPTİMİZASYON VE HESAPLAMA MOTORU
# ==============================================================================
if "results" not in st.session_state:
    st.session_state["results"] = None

if active_excel_source is not None:
    if st.button("🚀 Senaryoyu Hesapla ve Optimize Et", type="primary", key="btn_run", use_container_width=True):
        with st.spinner("Matematiksel kısıtlar, arızalar ve güncel siparişler hesaplanıyor..."):
            st.session_state["results"] = run_scheduler_pipeline(
                excel_source=active_excel_source,
                p6_debi=sim_p6_debi,
                kultur_suresi=sim_kultur_suresi,
                tank_cip_suresi=sim_tank_cip_suresi,
                max_kultur_bekleme=sim_max_kultur_bekleme,
                makine_max_calisma=sim_makine_max_calisma,
                p6_cip_limit=sim_p6_cip_limit,
                p6_cip_suresi=sim_p6_cip_suresi,
                gunluk_mesai_saati=sim_mesai_saati,
                opt_mode=sim_opt_mode,
                ariza_aktif=sim_ariza_aktif,
                ariza_gun=sim_ariza_gun,
                ariza_makine=sim_ariza_makine,
                ariza_saat_str=sim_ariza_saat_str,
                ariza_sure=sim_ariza_sure,
            )
        st.success("✅ Senaryo optimizasyonu başarıyla tamamlandı!")
else:
    st.warning("👈 Başlamak için sol menüden veri seçin ve hesaplamayı başlatın.")

if st.session_state["results"] is not None:
    results = st.session_state["results"]
    toplam_eksik_ton = results["toplam_eksik_genel"]

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Ortalama Günlük Üretim", f"{results['ort_gerceklesen']:.1f} T")
    col2.metric("P6 Efektif Hat Doygunluğu", f"%{results['genel_p6_oee']:.1f}")
    col3.metric("04:00 Hedef Uyum Oranı", f"%{results['genel_uyum']:.1f}")
    col4.metric("Karşılanamayan / Eksik Kalan", f"{toplam_eksik_ton:.1f} Ton", delta=f"{'-' if toplam_eksik_ton > 0 else ''}{toplam_eksik_ton:.1f} T", delta_color="inverse")

    st.download_button(
        label="📥 Nihai Excel Çizelgesini İndir (.xlsx)",
        data=results["excel_data"].getvalue(),
        file_name=f"Sutas_Uretim_Cizelgesi_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("---")

    # TAB YAPILANDIRMASI
    tab_options = ["📊 Yönetici Özeti"]
    
    if st.session_state["is_admin"]:
        tab_options.append("⚖️ Senaryo Kıyaslama (What-If)")

    tab_options.append("🔍 Denetim Logu")

    if st.session_state["is_admin"]:
        tab_options.extend(["📈 Darboğaz & Kapasite", "👥 İşgücü Analizi"])

    tab_options.extend(["📊 Gantt Şeması", "📅 Günlük Çizelgeler"])

    current_tab = st.radio("Görünüm Seçin:", tab_options, horizontal=True, key="selected_tab", label_visibility="collapsed")

    if current_tab == "📊 Yönetici Özeti":
        if toplam_eksik_ton > 0.05:
            st.error(f"🚨 **KRİTİK KAPASİTE UYARISI:** Bu hafta P6 pastörizatör ve tesis darboğazı nedeniyle **{toplam_eksik_ton:.1f} Ton** sipariş karşılanamadı / 04:00 mesai sınırına takıldı.")
        else:
            st.success("✅ **MÜKEMMEL OPERASYONEL PERFORMANS:** Bu haftaki tüm siparişler 04:00 mesai penceresi dolmadan %100 oranında eksiksiz karşılandı. Karşılanamayan sipariş: **0.0 Ton**")
        st.subheader("Haftalık & Günlük KPI Tablosu")
        st.dataframe(results["df_kpi"], use_container_width=True)
        st.pyplot(results["fig_kpi"])

    elif current_tab == "⚖️ Senaryo Kıyaslama (What-If)" and st.session_state["is_admin"]:
        st.subheader("⚖️ Stratejik Senaryo Kıyaslama ve Kapasite Analizi")
        st.markdown("Farklı operasyonel stratejilerin ve kapasite yatırımlarının tesis çıktısına etkisini yan yana kıyaslayın:")
        with st.spinner("Karşılaştırma senaryoları simüle ediliyor..."):
            res_curr = results
            res_max_p6 = run_scheduler_pipeline(
                excel_source=active_excel_source, p6_debi=18.0, kultur_suresi=sim_kultur_suresi,
                tank_cip_suresi=sim_tank_cip_suresi, max_kultur_bekleme=sim_max_kultur_bekleme, makine_max_calisma=sim_makine_max_calisma,
                p6_cip_limit=sim_p6_cip_limit, p6_cip_suresi=sim_p6_cip_suresi, gunluk_mesai_saati=sim_mesai_saati,
                opt_mode=sim_opt_mode, ariza_aktif=sim_ariza_aktif, ariza_gun=sim_ariza_gun, ariza_makine=sim_ariza_makine,
                ariza_saat_str=sim_ariza_saat_str, ariza_sure=sim_ariza_sure,
            )
            res_opt_cult = run_scheduler_pipeline(
                excel_source=active_excel_source, p6_debi=sim_p6_debi, kultur_suresi=1.0,
                tank_cip_suresi=sim_tank_cip_suresi, max_kultur_bekleme=sim_max_kultur_bekleme, makine_max_calisma=sim_makine_max_calisma,
                p6_cip_limit=sim_p6_cip_limit, p6_cip_suresi=sim_p6_cip_suresi, gunluk_mesai_saati=sim_mesai_saati,
                opt_mode=sim_opt_mode, ariza_aktif=sim_ariza_aktif, ariza_gun=sim_ariza_gun, ariza_makine=sim_ariza_makine,
                ariza_saat_str=sim_ariza_saat_str, ariza_sure=sim_ariza_sure,
            )
            res_both = run_scheduler_pipeline(
                excel_source=active_excel_source, p6_debi=18.0, kultur_suresi=1.0,
                tank_cip_suresi=sim_tank_cip_suresi, max_kultur_bekleme=sim_max_kultur_bekleme, makine_max_calisma=sim_makine_max_calisma,
                p6_cip_limit=sim_p6_cip_limit, p6_cip_suresi=sim_p6_cip_suresi, gunluk_mesai_saati=sim_mesai_saati,
                opt_mode=sim_opt_mode, ariza_aktif=sim_ariza_aktif, ariza_gun=sim_ariza_gun, ariza_makine=sim_ariza_makine,
                ariza_saat_str=sim_ariza_saat_str, ariza_sure=sim_ariza_sure,
            )

        comp_data = [
            {"Performans Göstergesi": "P6 Debi Hızı (Ton/Sa)", "1. Aktif Simülasyonun (Senin Kısıtların)": f"{res_curr['p6_debi']:.1f} T/Sa", "2. Maksimum P6 Önerisi (18 T/Sa)": f"{res_max_p6['p6_debi']:.1f} T/Sa", "3. Optimum Kültür Önerisi (1.0 Sa)": f"{res_opt_cult['p6_debi']:.1f} T/Sa", "4. Tam Entegre İkili İyileştirme": f"{res_both['p6_debi']:.1f} T/Sa"},
            {"Performans Göstergesi": "Mayalama Süresi (Saat)", "1. Aktif Simülasyonun (Senin Kısıtların)": f"{res_curr['kultur_suresi']:.2f} Sa", "2. Maksimum P6 Önerisi (18 T/Sa)": f"{res_max_p6['kultur_suresi']:.2f} Sa", "3. Optimum Kültür Önerisi (1.0 Sa)": f"{res_opt_cult['kultur_suresi']:.2f} Sa", "4. Tam Entegre İkili İyileştirme": f"{res_both['kultur_suresi']:.2f} Sa"},
            {"Performans Göstergesi": "Haftalık Gerçekleşen Tonaj", "1. Aktif Simülasyonun (Senin Kısıtların)": f"{res_curr['toplam_gerceklesen_genel']:.1f} Ton", "2. Maksimum P6 Önerisi (18 T/Sa)": f"{res_max_p6['toplam_gerceklesen_genel']:.1f} Ton", "3. Optimum Kültür Önerisi (1.0 Sa)": f"{res_opt_cult['toplam_gerceklesen_genel']:.1f} Ton", "4. Tam Entegre İkili İyileştirme": f"{res_both['toplam_gerceklesen_genel']:.1f} Ton"},
            {"Performans Göstergesi": "Karşılanamayan / Eksik Tonaj", "1. Aktif Simülasyonun (Senin Kısıtların)": f"{res_curr['toplam_eksik_genel']:.1f} Ton", "2. Maksimum P6 Önerisi (18 T/Sa)": f"{res_max_p6['toplam_eksik_genel']:.1f} Ton", "3. Optimum Kültür Önerisi (1.0 Sa)": f"{res_opt_cult['toplam_eksik_genel']:.1f} Ton", "4. Tam Entegre İkili İyileştirme": f"{res_both['toplam_eksik_genel']:.1f} Ton"},
            {"Performans Göstergesi": "04:00 Hedef Uyum Oranı (% OTIF)", "1. Aktif Simülasyonun (Senin Kısıtların)": f"%{res_curr['genel_uyum']:.1f}", "2. Maksimum P6 Önerisi (18 T/Sa)": f"%{res_max_p6['genel_uyum']:.1f}", "3. Optimum Kültür Önerisi (1.0 Sa)": f"%{res_opt_cult['genel_uyum']:.1f}", "4. Tam Entegre İkili İyileştirme": f"%{res_both['genel_uyum']:.1f}"},
            {"Performans Göstergesi": "P6 Efektif Hat Doygunluğu (%)", "1. Aktif Simülasyonun (Senin Kısıtların)": f"%{res_curr['genel_p6_oee']:.1f}", "2. Maksimum P6 Önerisi (18 T/Sa)": f"%{res_max_p6['genel_p6_oee']:.1f}", "3. Optimum Kültür Önerisi (1.0 Sa)": f"%{res_opt_cult['genel_p6_oee']:.1f}", "4. Tam Entegre İkili İyileştirme": f"%{res_both['genel_p6_oee']:.1f}"},
        ]
        st.dataframe(pd.DataFrame(comp_data), use_container_width=True)

        kurtarilan_p6_ton = max(0.0, res_curr['toplam_eksik_genel'] - res_max_p6['toplam_eksik_genel'])
        kurtarilan_both_ton = max(0.0, res_curr['toplam_eksik_genel'] - res_both['toplam_eksik_genel'])

        st.info(
            f"💡 **Yönetici & Kapasite Karar Notu:**\n"
            f"* **P6 Kapasite Artışı (18 T/Sa):** Aktif senaryona kıyasla haftalık **{kurtarilan_p6_ton:.1f} Ton** ek üretim sağlar.\n"
            f"* **Optimum Kültür Süresi (1.0 Sa / 60 dk):** Mayalama süresini 1 saate çekerek tank devir hızını artırır, gece hazırlığını rahatlatır ve hatların sabah 08:00'de kesintisiz doluma başlamasını garantiler.\n"
            f"* **Tam Entegre İkili İyileştirme:** Her iki iyileştirme birlikte devreye alındığında haftalık **{kurtarilan_both_ton:.1f} Ton** eksik sipariş kurtarılır."
        )

    elif current_tab == "🔍 Denetim Logu":
        st.subheader("P6 ve Tank Geçişleri Denetim Günlüğü")
        st.dataframe(results["df_audit"], use_container_width=True)

    elif current_tab == "📈 Darboğaz & Kapasite" and st.session_state["is_admin"]:
        st.subheader("Sistem Darboğazları & Kapasite Doluluk Oranları")
        st.pyplot(results["fig_db1"])
        st.pyplot(results["fig_hm"])

    elif current_tab == "👥 İşgücü Analizi" and st.session_state["is_admin"]:
        st.subheader("İşgücü İhtiyacı Seviyelendirme Analizi")
        isgucu_secenekleri = ["📊 Haftalık Genel Ortalama"] + list(results["gunluk_saatlik_isgucu"].keys())
        secilen_isgucu_gorunumu = st.selectbox("İşgücü Görünümü Seçin:", isgucu_secenekleri)

        mesai_h = results["mesai_h"]
        saatler = [f"{8+i:02d}:00" if 8+i < 24 else f"{8+i-24:02d}:00" for i in range(mesai_h)]

        if secilen_isgucu_gorunumu == "📊 Haftalık Genel Ortalama":
            toplam_saatlik = [0.0] * mesai_h
            for g_isim, g_vals in results["gunluk_saatlik_isgucu"].items():
                for h_i, val in enumerate(g_vals):
                    toplam_saatlik[h_i] += val
            gosterilecek_isgucu = [round(v / max(1, results["gun_sayisi"]), 1) for v in toplam_saatlik]
            grafik_baslik = "Haftalık Ortalama Saatlik İşgücü İhtiyacı (Kişi / Saat)"
        else:
            gosterilecek_isgucu = [round(v, 1) for v in results["gunluk_saatlik_isgucu"][secilen_isgucu_gorunumu]]
            grafik_baslik = f"{secilen_isgucu_gorunumu} Günü Saatlik İşgücü İhtiyacı (Kişi / Saat)"

        peak_val = max(gosterilecek_isgucu) if gosterilecek_isgucu else 0.0
        avg_val = round(sum(gosterilecek_isgucu) / max(1, len(gosterilecek_isgucu)), 1)
        gunduz_avg = round(sum(gosterilecek_isgucu[:10]) / 10.0, 1) if len(gosterilecek_isgucu) >= 10 else 0.0
        gece_avg = round(sum(gosterilecek_isgucu[10:]) / max(1, len(gosterilecek_isgucu[10:])), 1)

        c_ig1, c_ig2, c_ig3, c_ig4 = st.columns(4)
        c_ig1.metric("En Yüksek İhtiyaç (Peak)", f"{peak_val:.1f} Kişi")
        c_ig2.metric("Ortalama Vardiya Yükü", f"{avg_val:.1f} Kişi")
        c_ig3.metric("Gündüz (08:00 - 18:00)", f"{gunduz_avg:.1f} Kişi")
        c_ig4.metric("Gece (18:00 - 04:00)", f"{gece_avg:.1f} Kişi")

        fig_dinamik_ig, ax_dig = plt.subplots(figsize=(11.5, 3.4), dpi=200)
        fig_dinamik_ig.patch.set_facecolor("#FFFFFF")
        bars_dig = ax_dig.bar(range(mesai_h), gosterilecek_isgucu, color="#2E6BA8", alpha=0.85, edgecolor="#1F4E78", width=0.65)
        ax_dig.set_xticks(range(mesai_h))
        ax_dig.set_xticklabels(saatler, rotation=45, ha="right", fontsize=9, fontweight="bold")
        ax_dig.set_title(grafik_baslik, fontsize=11, fontweight="bold", pad=14)
        ax_dig.set_ylabel("Gereken İşgücü (Kişi)", fontsize=10, fontweight="bold")
        ax_dig.set_xlabel("Günün Saatleri (08:00 - 04:00 Mesai Penceresi)", fontsize=10, fontweight="bold", labelpad=8)
        ax_dig.grid(axis="y", linestyle="--", alpha=0.5)

        for bar in bars_dig:
            h = bar.get_height()
            if h > 0.0:
                ax_dig.text(bar.get_x() + bar.get_width()/2, h + 0.3, f"{h:.1f}", ha='center', va='bottom', fontsize=8.5, fontweight='bold')

        ax_dig.set_ylim(0, max(peak_val * 1.25, 5.0))
        plt.tight_layout()
        st.pyplot(fig_dinamik_ig)

    # --------------------------------------------------------------------------
    # İNTERAKTİF GANTT ŞEMASI (PLOTLY)
    # --------------------------------------------------------------------------
    elif current_tab == "📊 Gantt Şeması":
        st.subheader("Tesis İçi Günlük Üretim & Zaman Çizelgesi (İnteraktif Gantt)")
        df_gantt_all = pd.DataFrame(results["all_schedule_rows"])
        if not df_gantt_all.empty:
            gun_isimleri_gantt = list(dict.fromkeys(df_gantt_all["gun_adi"].tolist()))
            secilen_gantt_gun = st.selectbox("Gantt Şemasını Görüntülemek İstediğiniz Günü Seçin:", gun_isimleri_gantt, key="gantt_day_sel")
            df_gantt_filtered = df_gantt_all[df_gantt_all["gun_adi"] == secilen_gantt_gun].copy()

            g_tonaj = df_gantt_filtered[df_gantt_filtered["Süt Tipi"] != "DURUŞ"]["Miktar (Ton)"].sum()
            g_parti = len(df_gantt_filtered[df_gantt_filtered["Süt Tipi"] != "DURUŞ"])
            g_ariza = len(df_gantt_filtered[df_gantt_filtered["Süt Tipi"] == "DURUŞ"])

            cg1, cg2, cg3, cg4 = st.columns(4)
            cg1.metric("Toplam Gerçekleşen Üretim", f"{g_tonaj:.1f} Ton")
            cg2.metric("Üretilen Parti Sayısı", f"{g_parti} Parti")
            cg3.metric("Aktif Çalışan Hat Sayısı", f"{df_gantt_filtered['Makine'].nunique()} Hat")
            cg4.metric("Duruş / Arıza Durumu", f"{'1 Kesinti ⚠️' if g_ariza > 0 else 'Kesintisiz 🟢'}")

            color_map = {
                "TAM YAĞLI": "#1F4E78",
                "YARIM YAĞLI": "#2E75B6",
                "%5 YAĞLI": "#548235",
                "PAKSÜT": "#C55A11",
                "DURUŞ": "#C00000",
            }

            fig_plotly = px.timeline(
                df_gantt_filtered,
                x_start="dt_start",
                x_end="dt_end",
                y="Makine",
                color="Süt Tipi",
                color_discrete_map=color_map,
                category_orders={"Makine": MAKINE_LISTESI},
                hover_data={
                    "Sipariş ID": True,
                    "Ürün Adı": True,
                    "Miktar (Ton)": True,
                    "Tahsis Tank": True,
                    "Kalıp/Gramaj": True,
                    "Hız (T/Sa)": True,
                    "Kültür & CIP Hijyen Notu": True,
                    "dt_start": False,
                    "dt_end": False,
                },
                title=f"{secilen_gantt_gun} Günü İnteraktif Üretim Akışı (Mouse ile üzerine gelin, yakınlaştırın)",
            )

            fig_plotly.update_yaxes(autorange="reversed")
            fig_plotly.update_layout(
                xaxis=dict(
                    title="Vardiya Saatleri",
                    tickformat="%H:%M",
                    dtick=3600000,
                    showgrid=True,
                    gridcolor="#E0E4E8",
                ),
                yaxis=dict(title="Makineler", showgrid=True, gridcolor="#F0F2F6"),
                plot_bgcolor="#FFFFFF",
                paper_bgcolor="#FAFAFC",
                height=450,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                margin=dict(l=40, r=40, t=60, b=40),
            )

            st.plotly_chart(fig_plotly, use_container_width=True)
        else:
            st.info("Gantt şeması oluşturmak için lütfen sol menüden simülasyonu çalıştırın.")

    elif current_tab == "📅 Günlük Çizelgeler":
        st.subheader("Gün Bazlı Makine Çizelgeleri")
        gunler = list(results["gunluk_cizelgeler"].keys())
        selected_day = st.selectbox("Görüntülenecek Günü Seçin", gunler, key="day_selector")
        if selected_day:
            df_to_show = results["gunluk_cizelgeler"][selected_day]
            display_df = df_to_show.drop(columns=["dt_start", "dt_end", "gun_adi"], errors="ignore")
            st.dataframe(display_df, use_container_width=True)

# ------------------------------------------------------------------------------
# SAYFA EN ALTI: SİLİNEMEZ TELİF VE GELİŞTİRİCİ FOOTER'I
# ------------------------------------------------------------------------------
st.markdown("---")
st.markdown(
    f"""
    <div style="text-align: center; color: #595959; font-size: 13px; padding: 15px 0;">
        <b>Sütaş Karacabey Master Scheduler & Decision Support System (DSS)</b><br>
        Developer: <b>{DEVELOPER_NAME}</b><br>
        <span style="font-size: 11px; color: #8C8C8C;">© 2026 Tüm Hakları Saklıdır.</span>
    </div>
    """,
    unsafe_allow_html=True,
)
